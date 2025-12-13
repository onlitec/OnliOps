from openpyxl import load_workbook

wb = load_workbook('Rede_VLANs_Condominio_Calabasas.xlsx')
ws = wb['IP - CFTV (VLAN 40)']

print('=' * 80)
print('VERIFICAÇÃO COMPLETA DA PLANILHA CFTV')
print('=' * 80)

print(f'\nTotal de linhas: {ws.max_row}')
print(f'Total de colunas: {ws.max_column}')

print('\nTítulo (linha 1):')
print(f'  {ws["A1"].value}')

print('\nCabeçalho primeira tabela (linha 3):')
for i in range(1, 11):
    cell = ws.cell(3, i)
    rgb = cell.fill.start_color.rgb if cell.fill.start_color else None
    print(f'  Col {i}: {cell.value} | RGB: {rgb}')

print('\nPrimeiras 5 linhas de dados:')
for row in range(4, min(9, ws.max_row + 1)):
    print(f'  Linha {row}: {[ws.cell(row, i).value for i in range(1, 6)]}')

print('\nProcurando segunda tabela (cabeçalho verde):')
found_green = False
for row in range(1, min(ws.max_row + 1, 100)):
    cell = ws.cell(row, 1)
    if cell.fill and cell.fill.start_color:
        rgb = str(cell.fill.start_color.rgb)
        if '00B050' in rgb or 'B050' in rgb:
            found_green = True
            print(f'  Linha {row} tem cabeçalho verde!')
            print(f'    Conteúdo: {[ws.cell(row, i).value for i in range(1, 6)]}')
            break

if not found_green:
    print('  Segunda tabela NÃO encontrada!')

print('\nVerificando senhas:')
senhas = []
for row in range(4, min(ws.max_row + 1, 30)):
    senha = ws.cell(row, 10).value
    if senha:
        senhas.append(senha)

senhas_unicas = list(set(senhas))
print(f'  Senhas encontradas: {senhas_unicas}')
if 'Hical@20#25' in senhas_unicas:
    print('  ✓ Senha atualizada corretamente!')
else:
    print('  ✗ Senha NÃO está atualizada!')

