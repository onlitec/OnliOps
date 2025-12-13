import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Criar uma planilha Excel simplificada com as melhorias identificadas

def create_updated_vlan_spreadsheet():
    # Criar workbook
    wb = Workbook()
    
    # Remover a planilha padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Cores atualizadas
    COLORS = {
        'header': '2F4F4F',  # Verde-azulado escuro para cabeçalhos
        'header_text': 'FFFFFF',  # Branco para texto de cabeçalho
        'vlan_10': 'E2EFDA',  # Verde muito claro - Management
        'vlan_20': 'DEEBF7',  # Azul muito claro - Data
        'vlan_30': 'FFF2CC',  # Amarelo muito claro - Voice
        'vlan_40': 'FCE4D6',  # Laranja muito claro - CFTV
        'vlan_50': 'E1D5E7',  # Roxo muito claro - Access Control
        'vlan_60': 'F4F4F4',  # Cinza muito claro - IoT
        'vlan_100': 'D9D9D9',  # Cinza claro - Guest
        'info': 'FFF9E6',  # Bege claro para informações
        'border': '000000',  # Preto para bordas
        'warning': 'FFE699',  # Amarelo para avisos
        'success': 'C6EFCE',  # Verde claro para sucesso
        'error': 'FFC7CE',    # Vermelho claro para erros
        'security': 'FF9999', # Vermelho para alertas de segurança
    }
    
    # Criar planilha de resumo geral
    create_summary_sheet(wb, COLORS)
    
    # Criar planilha de VLANs detalhadas
    create_detailed_vlan_sheet(wb, COLORS)
    
    # Criar planilha de equipamentos
    create_equipment_sheet(wb, COLORS)
    
    # Criar planilha de endereçamento IP
    create_ip_addressing_sheet(wb, COLORS)
    
    # Salvar arquivo
    wb.save('Projeto_Logico_Rede_Calabasas_Atualizado.xlsx')
    print("Planilha atualizada gerada com sucesso!")

def apply_border(cell):
    """Aplica borda fina à célula"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

def create_summary_sheet(wb, COLORS):
    """Cria a planilha de resumo geral"""
    ws = wb.create_sheet('Resumo Geral', 0)
    
    # Título principal
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'ARQUITETURA DE REDE COM SEGMENTAÇÃO POR VLANS - CONDOMÍNIO CALABASAS'
    title_cell.font = Font(bold=True, size=16, color='FFFFFF')
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Informações gerais
    ws.merge_cells('A2:H2')
    info_cell = ws['A2']
    info_cell.value = f'Documento gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Servidor HikCentral: 10.10.20.10 (VLAN 20)'
    info_cell.font = Font(size=10, italic=True)
    info_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(info_cell)
    ws.row_dimensions[2].height = 20
    
    # Cabeçalho da tabela
    headers = ['VLAN ID', 'Nome', 'Sub-rede', 'Gateway', 'Máscara', 'Descrição', 'Faixa de IPs', 'Cor']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 40
    
    # Dados das VLANs
    vlan_data = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', '255.255.255.0', 'Gerenciamento de infraestrutura', '10.10.10.2-254', 'Verde claro'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', '255.255.255.0', 'Rede corporativa (HikCentral)', '10.10.20.2-254', 'Azul claro'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', '255.255.255.0', 'Telefonia IP', '10.10.30.2-254', 'Amarelo claro'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', '255.255.255.0', 'Sistema de CFTV', '10.10.40.2-254', 'Laranja claro'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', '255.255.255.0', 'Controle de acesso', '10.10.50.2-254', 'Roxo claro'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', '255.255.255.0', 'Dispositivos IoT', '10.10.60.2-254', 'Cinza claro'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', '255.255.255.0', 'WiFi visitantes', '10.10.100.2-254', 'Cinza']
    ]
    
    # Cores das VLANs
    vlan_colors = {
        10: COLORS['vlan_10'],
        20: COLORS['vlan_20'],
        30: COLORS['vlan_30'],
        40: COLORS['vlan_40'],
        50: COLORS['vlan_50'],
        60: COLORS['vlan_60'],
        100: COLORS['vlan_100']
    }
    
    for row_idx, row_data in enumerate(vlan_data, start=4):
        vlan_id = row_data[0]
        color = vlan_colors[vlan_id]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 25
    
    # Legenda e informações adicionais
    start_row = len(vlan_data) + 6
    ws.merge_cells(f'A{start_row}:H{start_row}')
    legend_title = ws.cell(row=start_row, column=1)
    legend_title.value = 'LEGENDA E INFORMAÇÕES IMPORTANTES'
    legend_title.font = Font(bold=True, size=12, color='FFFFFF')
    legend_title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    legend_title.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(legend_title)
    ws.row_dimensions[start_row].height = 25
    
    legend_items = [
        ['SERVIDOR HIKCENTRAL:', 'IP: 10.10.20.10 na VLAN 20 (Data). Gerencia CFTV (VLAN 40) e Controle de Acesso (VLAN 50)'],
        ['CORES:', 'Cada VLAN possui uma cor específica para facilitar identificação visual nas planilhas'],
        ['IPs ESTÁTICOS:', 'Todos os equipamentos de CFTV e Controle de Acesso devem usar IPs estáticos'],
        ['SEGURANÇA:', 'Credenciais padrão devem ser alteradas imediatamente após a migração'],
        ['BACKUP:', 'Manter cópias de segurança da configuração de todos os equipamentos'],
        ['MONITORAMENTO:', 'Implementar monitoramento proativo de conectividade e desempenho']
    ]
    
    for idx, (label, desc) in enumerate(legend_items, start=start_row + 1):
        ws.merge_cells(f'A{idx}:B{idx}')
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        apply_border(label_cell)
        
        ws.merge_cells(f'C{idx}:H{idx}')
        desc_cell = ws.cell(row=idx, column=3, value=desc)
        desc_cell.font = Font(size=10)
        desc_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        apply_border(desc_cell)
        ws.row_dimensions[idx].height = 20
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 18, 15, 15, 35, 25, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_detailed_vlan_sheet(wb, COLORS):
    """Cria a planilha de VLANs detalhadas"""
    ws = wb.create_sheet('VLANs Detalhadas', 1)
    
    # Título
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'DETALHAMENTO DAS VLANS'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Cabeçalhos
    headers = ['VLAN ID', 'Nome', 'Sub-rede', 'Gateway', 'Equipamentos', 'Responsável', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_border(cell)
    ws.row_dimensions[2].height = 25
    
    # Dados das VLANs
    vlan_details = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', 'Switches, Roteadores, NVRs', 'TI Condomínio', 'Ativa'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', 'Computadores, Servidores, Impressoras', 'TI Condomínio', 'Ativa'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', 'Telefones IP, PBX', 'TI Condomínio', 'Ativa'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', 'Câmeras, NVRs, Conversores', 'Segurança', 'Ativa'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', 'Leitores, Controladores, Câmeras', 'Segurança', 'Ativa'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', 'Sensores, Dispositivos Inteligentes', 'TI Condomínio', 'Planejada'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', 'WiFi Visitantes', 'TI Condomínio', 'Ativa']
    ]
    
    # Cores das VLANs
    vlan_colors = {
        10: COLORS['vlan_10'],
        20: COLORS['vlan_20'],
        30: COLORS['vlan_30'],
        40: COLORS['vlan_40'],
        50: COLORS['vlan_50'],
        60: COLORS['vlan_60'],
        100: COLORS['vlan_100']
    }
    
    for row_idx, row_data in enumerate(vlan_details, start=3):
        vlan_id = row_data[0]
        color = vlan_colors[vlan_id]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 20
    
    # Ajustar largura das colunas
    column_widths = [10, 20, 15, 15, 30, 15, 10]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_equipment_sheet(wb, COLORS):
    """Cria a planilha de equipamentos"""
    ws = wb.create_sheet('Equipamentos', 2)
    
    # Título
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = 'INVENTÁRIO DE EQUIPAMENTOS DE GERENCIAMENTO'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Cabeçalhos
    headers = ['VLAN', 'IP', 'TAG', 'Tipo', 'Modelo', 'Localização']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_border(cell)
    ws.row_dimensions[2].height = 25
    
    # Equipamentos de gerenciamento
    equipments = [
        ['VLAN 10', '10.10.10.1', 'ROTEADOR-01', 'Roteador', 'MIKROTIK RB750Gr3', 'Rack Principal'],
        ['VLAN 10', '10.10.10.2', 'SW-00-RK-00', 'Switch', 'DS-3E1526P-EI', 'Rack Principal'],
        ['VLAN 10', '10.10.10.3', 'SW-01-TR-01', 'Switch', 'DS-3E1318P-EI/M', 'Torre 1'],
        ['VLAN 10', '10.10.10.4', 'SW-02-TR-02', 'Switch', 'DS-3E1526P-EI', 'Torre 2'],
        ['VLAN 10', '10.10.10.5', 'SW-03-TR-03', 'Switch', 'DS-3E1526P-EI', 'Torre 3'],
        ['VLAN 10', '10.10.10.6', 'SW-04-TR-04', 'Switch', 'DS-3E1526P-EI', 'Torre 4'],
        ['VLAN 10', '10.10.10.7', 'SW-01-CP-01', 'Switch', 'DS-3E1309P-EI', 'Caixa Perimetral 1'],
        ['VLAN 10', '10.10.10.8', 'SW-02-CP-02', 'Switch', 'DS-3E1309P-EI', 'Caixa Perimetral 2'],
        ['VLAN 10', '10.10.10.9', 'SW-03-CP-03', 'Switch', 'DS-3E1309P-EI', 'Caixa Perimetral 3'],
        ['VLAN 10', '10.10.10.10', 'SW-04-CP-04', 'Switch', 'DS-3E1309P-EI', 'Caixa Perimetral 4'],
        ['VLAN 20', '10.10.20.10', 'SERVIDOR-01', 'Servidor', 'HikCentral', 'Rack Principal']
    ]
    
    for row_idx, row_data in enumerate(equipments, start=3):
        vlan = row_data[0]
        # Determinar cor com base na VLAN
        if '10' in vlan:
            color = COLORS['vlan_10']
        elif '20' in vlan:
            color = COLORS['vlan_20']
        else:
            color = 'FFFFFF'
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 20
    
    # Ajustar largura das colunas
    column_widths = [10, 15, 15, 15, 20, 25]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_ip_addressing_sheet(wb, COLORS):
    """Cria a planilha de endereçamento IP"""
    ws = wb.create_sheet('Endereçamento IP', 3)
    
    # Título
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'DETALHAMENTO DO ENDEREÇAMENTO IP POR VLAN'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Informações gerais
    ws.merge_cells('A2:G2')
    info_cell = ws['A2']
    info_cell.value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    info_cell.font = Font(size=10, italic=True)
    info_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(info_cell)
    ws.row_dimensions[2].height = 20
    
    # Cabeçalhos
    headers = ['VLAN ID', 'Nome VLAN', 'Sub-rede', 'Gateway', 'Faixa Utilizável', 'IPs Reservados', 'Observações']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 40
    
    # Dados das VLANs
    vlan_details = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', '10.10.10.2-254', '10.10.10.101-254', 'Switches: 10.10.10.2-50, NVRs: 10.10.10.51-100'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', '10.10.20.2-254', '10.10.20.201-254', 'Servidores: 10.10.20.2-50, Workstations: 10.10.20.51-200'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', '10.10.30.2-254', '10.10.30.201-254', 'Telefones IP: 10.10.30.11-200'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', '10.10.40.2-254', '10.10.40.241-254', 'Câmeras: 10.10.40.51-200, NVRs: 10.10.40.2-10'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', '10.10.50.2-254', '10.10.50.241-254', 'Leitores: 10.10.50.101-200, Controladores: 10.10.50.201-240'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', '10.10.60.2-254', '10.10.60.201-254', 'Dispositivos IoT'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', '10.10.100.2-254', '10.10.100.201-254', 'WiFi Visitantes']
    ]
    
    # Cores das VLANs
    vlan_colors = {
        10: COLORS['vlan_10'],
        20: COLORS['vlan_20'],
        30: COLORS['vlan_30'],
        40: COLORS['vlan_40'],
        50: COLORS['vlan_50'],
        60: COLORS['vlan_60'],
        100: COLORS['vlan_100']
    }
    
    for row_idx, row_data in enumerate(vlan_details, start=4):
        vlan_id = row_data[0]
        color = vlan_colors[vlan_id]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 25
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 18, 15, 20, 15, 30]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

# Executar a função principal
if __name__ == "__main__":
    create_updated_vlan_spreadsheet()