import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime

# Cores baseadas na imagem
COLORS = {
    'header_red': 'FF0000',  # Vermelho para cabeçalho primeira tabela
    'header_green': '00B050',  # Verde para cabeçalho segunda tabela
    'row_white': 'FFFFFF',  # Branco
    'row_green': 'E2EFDA',  # Verde claro alternado
    'border': '000000',  # Preto
    'text_dark': '000000',  # Preto
    'title_bg': 'FFFFFF',  # Branco para fundo do título
    'vlan_40': 'FCE4D6',  # Laranja claro - CFTV
    'vlan_50': 'E1D5E7',  # Roxo claro - Access Control
}

# Senha atualizada
SENHA_HI = 'Hical@20#25'
SENHA_OPERADOR = 'cc2025'

# Contadores de IP
ip_counters = {
    10: {'switches': 2, 'nvrs': 51, 'others': 100},
    40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
    50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
}

def extract_ip(ip_str):
    if pd.isna(ip_str):
        return None
    ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    match = re.search(ip_pattern, str(ip_str))
    return match.group(0) if match else None

def calculate_new_ip(old_ip, vlan_id, equipment_type=None):
    global ip_counters
    if not old_ip:
        return None
    try:
        last_octet = int(old_ip.split('.')[-1])
        if vlan_id == 10:
            if 'SW' in str(equipment_type).upper() or 'ROTEADOR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['switches']
                ip_counters[10]['switches'] = min(ip_counters[10]['switches'] + 1, 50)
            elif 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['nvrs']
                ip_counters[10]['nvrs'] = min(ip_counters[10]['nvrs'] + 1, 100)
            else:
                new_octet = ip_counters[10]['others']
                ip_counters[10]['others'] = min(ip_counters[10]['others'] + 1, 254)
            return f'10.10.10.{new_octet}'
        elif vlan_id == 40:
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['nvrs']
                ip_counters[40]['nvrs'] = min(ip_counters[40]['nvrs'] + 1, 10)
            elif 'CM-' in str(equipment_type).upper() or 'CONVERSOR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['converters']
                ip_counters[40]['converters'] = min(ip_counters[40]['converters'] + 1, 240)
            else:
                new_octet = ip_counters[40]['cameras']
                ip_counters[40]['cameras'] = min(ip_counters[40]['cameras'] + 1, 200)
            return f'10.10.40.{new_octet}'
        elif vlan_id == 50:
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['nvrs']
                ip_counters[50]['nvrs'] = min(ip_counters[50]['nvrs'] + 1, 10)
            elif 'EL-' in str(equipment_type).upper() or 'ELEVADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['cameras_elevator']
                ip_counters[50]['cameras_elevator'] = min(ip_counters[50]['cameras_elevator'] + 1, 50)
            elif 'LF-' in str(equipment_type).upper() or 'LEITOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['readers']
                ip_counters[50]['readers'] = min(ip_counters[50]['readers'] + 1, 200)
            elif 'CE-' in str(equipment_type).upper() or 'CONTROLADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['controllers']
                ip_counters[50]['controllers'] = min(ip_counters[50]['controllers'] + 1, 240)
            else:
                new_octet = ip_counters[50]['cameras_portaria']
                ip_counters[50]['cameras_portaria'] = min(ip_counters[50]['cameras_portaria'] + 1, 100)
            return f'10.10.50.{new_octet}'
        else:
            base = vlan_id * 10 if vlan_id < 100 else 100
            return f'10.10.{base}.{last_octet}'
    except:
        return None

def get_vlan_for_equipment(tag, modelo, localizacao, ip_atual, sheet_type='CFTV'):
    """Determina VLAN baseado no tipo de equipamento e contexto"""
    if pd.isna(tag):
        tag = ''
    if pd.isna(modelo):
        modelo = ''
    if pd.isna(localizacao):
        localizacao = ''
    
    tag_str = str(tag).upper()
    modelo_str = str(modelo).upper()
    localizacao_str = str(localizacao).upper()
    
    # Switches e roteadores - sempre VLAN 10
    if any(x in tag_str for x in ['SW-', 'RK-', 'ROTEADOR']):
        return 10
    
    # NVRs - VLAN 10 para gerenciamento
    if 'NVR' in tag_str or 'NVR' in modelo_str:
        return 10
    
    # Se está na planilha de CFTV, priorizar VLAN 40
    if sheet_type == 'CFTV':
        # Câmeras CFTV (CA-) - sempre VLAN 40 independente da localização
        if 'CA-' in tag_str:
            return 40
        # Conversores de mídia
        if 'CM-' in tag_str or 'CONVERSOR' in modelo_str:
            return 40
        # Por IP - se está em 10.10.0.x, é CFTV
        if ip_atual and ip_atual.startswith('10.10.0.'):
            return 40
        # Default para CFTV
        return 40
    
    # Se está na planilha de Controle de Acesso, priorizar VLAN 50
    if sheet_type == 'ACCESS':
        # Câmeras de elevador
        if 'EL-' in tag_str or 'ELEVADOR' in localizacao_str:
            return 50
        # LPR
        if 'LPR' in tag_str or 'LPR' in localizacao_str:
            return 50
        # Leitores biométricos
        if 'LF-' in tag_str or 'LEITOR' in modelo_str:
            return 50
        # Controladores
        if 'CE-' in tag_str or 'CONTROLADOR' in modelo_str:
            return 50
        # Portaria
        if 'PORTARIA' in localizacao_str or 'PORTÃO' in localizacao_str:
            return 50
        # Por IP - se está em 10.10.1.x, é Controle de Acesso
        if ip_atual and ip_atual.startswith('10.10.1.'):
            return 50
        # Default para Access Control
        return 50
    
    # Fallback baseado em IP
    if ip_atual and ip_atual.startswith('10.10.0.'):
        return 40
    if ip_atual and ip_atual.startswith('10.10.1.'):
        return 50
    
    return 40

def apply_border(cell):
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )
    cell.border = thin_border

def create_cftv_sheet(wb, df, vlan_id=40):
    """Cria planilha CFTV com layout da imagem mas com todas as informações de VLAN"""
    ws = wb.create_sheet('IP - CFTV')
    
    # Configurar largura das colunas
    column_widths = [8, 12, 15, 15, 15, 20, 25, 40, 15, 18, 15, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Altura da linha do título
    ws.row_dimensions[1].height = 60
    
    # Título principal (linha 1) - estilo da imagem
    ws.merge_cells('A1:N1')
    title = ws['A1']
    title.value = 'LISTA DE EQUIPAMENTOS - CONDOMINIO CALABASAS'
    title.font = Font(bold=True, size=16, color=COLORS['text_dark'])
    title.fill = PatternFill(start_color=COLORS['title_bg'], end_color=COLORS['title_bg'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar logos (se existirem)
    try:
        # Logo HI (lado esquerdo)
        logo_paths = ['logo_hi.png', 'logo_hi.jpg', 'logo_HI.png', 'logo_HI.jpg']
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                img_hi = Image(logo_path)
                img_hi.width = 100
                img_hi.height = 60
                ws.add_image(img_hi, 'A1')
                break
        
        # Logo Calabasas (lado direito)
        calabasas_paths = ['logo_calabasas.png', 'logo_calabasas.jpg', 'Calabasas.png', 'Calabasas.jpg']
        for logo_path in calabasas_paths:
            if os.path.exists(logo_path):
                img_cal = Image(logo_path)
                img_cal.width = 150
                img_cal.height = 60
                ws.add_image(img_cal, 'M1')
                break
    except Exception as e:
        print(f"Nota: Logos não encontrados ou erro ao adicionar: {e}")
    
    # Encontrar cabeçalho na planilha original
    header_row = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x) for x in row.values if pd.notna(x)]).upper()
        if 'ORDEM' in row_str and 'IP' in row_str:
            header_row = idx
            break
    if header_row is None:
        header_row = 4
    
    # Identificar colunas
    ordem_col = None
    tag_rack_col = None
    nomenclatura_col = None
    modelo_col = None
    localizacao_col = None
    ip_col = None
    usuario_op_col = None
    senha_op_col = None
    usuario_hi_col = None
    senha_hi_col = None
    
    for col_idx in range(len(df.columns)):
        col_header = str(df.iloc[header_row, col_idx]).upper()
        if 'ORDEM' in col_header and ordem_col is None:
            ordem_col = col_idx
        if 'TAG DO RACK' in col_header or 'TAG DA CAIXA' in col_header:
            tag_rack_col = col_idx
        if 'NOMENCLATURA' in col_header:
            nomenclatura_col = col_idx
        if 'MODELO' in col_header:
            modelo_col = col_idx
        if 'LOCALIZA' in col_header:
            localizacao_col = col_idx
        if 'IP' in col_header and ip_col is None and 'EQUIPAMENTO' not in col_header:
            ip_col = col_idx
        if 'USUARIO OPERADOR' in col_header or 'USUÁRIO OPERADOR' in col_header:
            usuario_op_col = col_idx
        if 'SENHA OPERADOR' in col_header:
            senha_op_col = col_idx
        if 'USUARIO HI' in col_header or 'USUÁRIO HI' in col_header:
            usuario_hi_col = col_idx
        if 'SENHA HI' in col_header:
            senha_hi_col = col_idx
    
    # Função para validar linha
    def is_valid_data_row(row_data):
        if all(pd.isna(x) or str(x).strip() == '' for x in row_data):
            return False
        row_str = ' '.join([str(x) for x in row_data if pd.notna(x)]).upper()
        header_keywords = ['ORDEM', 'TAG', 'IP', 'MODELO', 'LOCALIZA', 'USUARIO', 'SENHA', 'NOMENCLATURA']
        keyword_count = sum(1 for keyword in header_keywords if keyword in row_str)
        if keyword_count >= 3:
            if ordem_col is not None and ordem_col < len(row_data):
                ordem_val = row_data[ordem_col]
                ordem_str = str(ordem_val).strip() if pd.notna(ordem_val) else ''
                if not ordem_str.isdigit():
                    return False
        if ip_col is not None and ip_col < len(row_data):
            ip_val = row_data[ip_col]
            ip_extracted = extract_ip(ip_val)
            if ip_extracted:
                return True
        if ordem_col is not None and ordem_col < len(row_data):
            ordem_val = row_data[ordem_col]
            ordem_str = str(ordem_val).strip() if pd.notna(ordem_val) else ''
            if ordem_str.isdigit():
                return True
        return False
    
    # Primeira tabela - Cabeçalho vermelho (com informações de VLAN)
    start_row = 3
    current_row = start_row
    
    # Cabeçalho primeira tabela (vermelho) - INCLUINDO colunas de VLAN
    headers_table1 = ['ORDEM', 'TAG DO RACK', 'NOMENCLATURA\nTAG DA CÂMERA', 'MODELO', 
                      'LOCALIZAÇÃO CÂMERA', 'VLAN ID', 'VLAN Nome', 'IP ANTIGO', 'IP NOVO', 
                      'STATUS', 'USUARIO OPERADOR', 'SENHA OPERADOR', 'USUARIO HI', 'SENHA HI']
    
    for col_idx, header in enumerate(headers_table1, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header_red'], end_color=COLORS['header_red'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[current_row].height = 50
    current_row += 1
    
    # Dados primeira tabela (primeiras 20 linhas)
    count_table1 = 0
    for idx in range(header_row + 1, len(df)):
        if count_table1 >= 20:
            break
        
        row_data = df.iloc[idx].values
        
        if not is_valid_data_row(row_data):
            continue
        
        ordem_val = row_data[ordem_col] if ordem_col is not None else None
        ip_val = row_data[ip_col] if ip_col is not None else None
        
        if pd.isna(ordem_val) or not str(ordem_val).strip().isdigit():
            continue
        if not extract_ip(ip_val):
            continue
        
        # Extrair dados
        ordem = int(ordem_val) if pd.notna(ordem_val) else None
        tag_rack = row_data[tag_rack_col] if tag_rack_col is not None else None
        nomenclatura = row_data[nomenclatura_col] if nomenclatura_col is not None else None
        modelo = row_data[modelo_col] if modelo_col is not None else None
        localizacao = row_data[localizacao_col] if localizacao_col is not None else None
        ip_antigo = extract_ip(ip_val)
        
        # Determinar VLAN (contexto CFTV)
        vlan_id = get_vlan_for_equipment(tag_rack, modelo, localizacao, ip_antigo, 'CFTV')
        vlan_name = 'CFTV' if vlan_id == 40 else 'Access Control' if vlan_id == 50 else 'Management'
        
        # Calcular novo IP
        ip_novo = calculate_new_ip(ip_antigo, vlan_id, tag_rack)
        status = 'Migrado' if ip_novo else 'Pendente'
        
        # Credenciais
        usuario_op = row_data[usuario_op_col] if usuario_op_col is not None and pd.notna(row_data[usuario_op_col]) else 'operador'
        senha_op = row_data[senha_op_col] if senha_op_col is not None and pd.notna(row_data[senha_op_col]) else SENHA_OPERADOR
        usuario_hi = row_data[usuario_hi_col] if usuario_hi_col is not None and pd.notna(row_data[usuario_hi_col]) else 'admin'
        senha_hi = SENHA_HI  # Sempre usar senha atualizada
        
        # Criar linha com TODAS as informações
        row_values = [
            ordem, tag_rack, nomenclatura, modelo, localizacao,
            vlan_id, vlan_name, ip_antigo, ip_novo, status,
            usuario_op, senha_op, usuario_hi, senha_hi
        ]
        
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            # Linhas alternadas
            if count_table1 % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS['row_white'], end_color=COLORS['row_white'], fill_type='solid')
            else:
                cell.fill = PatternFill(start_color=COLORS['row_green'], end_color=COLORS['row_green'], fill_type='solid')
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        
        current_row += 1
        count_table1 += 1
    
    # Espaço entre tabelas
    current_row += 2
    
    # Segunda tabela - Cabeçalho verde (com informações de VLAN)
    headers_table2 = ['ORDEM', 'TAG DA CAIXA', 'NOMENCLATURA\nTAG DA CÂMERA', 'MODELO', 
                     'LOCALIZAÇÃO CÂMERA', 'VLAN ID', 'VLAN Nome', 'IP ANTIGO', 'IP NOVO',
                     'STATUS', 'USUARIO OPERADOR', 'SENHA OPERADOR', 'USUARIO HI', 'SENHA HI']
    
    for col_idx, header in enumerate(headers_table2, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header_green'], end_color=COLORS['header_green'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[current_row].height = 50
    current_row += 1
    
    # Dados segunda tabela
    count_table2 = 0
    ordem_counter = 1
    for idx in range(header_row + 1, len(df)):
        row_data = df.iloc[idx].values
        
        # Pular itens já usados na primeira tabela
        ordem_val = row_data[ordem_col] if ordem_col is not None else None
        if pd.notna(ordem_val) and str(ordem_val).strip().isdigit():
            ordem_num = int(ordem_val)
            if ordem_num <= 20:
                continue
        
        if not is_valid_data_row(row_data):
            continue
        
        ip_val = row_data[ip_col] if ip_col is not None else None
        if not extract_ip(ip_val):
            continue
        
        # Extrair dados
        tag_caixa = row_data[tag_rack_col] if tag_rack_col is not None else None
        nomenclatura = row_data[nomenclatura_col] if nomenclatura_col is not None else None
        modelo = row_data[modelo_col] if modelo_col is not None else None
        localizacao = row_data[localizacao_col] if localizacao_col is not None else None
        ip_antigo = extract_ip(ip_val)
        
        # Determinar VLAN (contexto CFTV)
        vlan_id = get_vlan_for_equipment(tag_caixa, modelo, localizacao, ip_antigo, 'CFTV')
        vlan_name = 'CFTV' if vlan_id == 40 else 'Access Control' if vlan_id == 50 else 'Management'
        
        # Calcular novo IP
        ip_novo = calculate_new_ip(ip_antigo, vlan_id, tag_caixa)
        status = 'Migrado' if ip_novo else 'Pendente'
        
        # Credenciais
        usuario_op = row_data[usuario_op_col] if usuario_op_col is not None and pd.notna(row_data[usuario_op_col]) else 'operador'
        senha_op = row_data[senha_op_col] if senha_op_col is not None and pd.notna(row_data[senha_op_col]) else SENHA_OPERADOR
        usuario_hi = row_data[usuario_hi_col] if usuario_hi_col is not None and pd.notna(row_data[usuario_hi_col]) else 'admin'
        senha_hi = SENHA_HI
        
        row_values = [
            ordem_counter, tag_caixa, nomenclatura, modelo, localizacao,
            vlan_id, vlan_name, ip_antigo, ip_novo, status,
            usuario_op, senha_op, usuario_hi, senha_hi
        ]
        
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            # Linhas alternadas
            if count_table2 % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS['row_white'], end_color=COLORS['row_white'], fill_type='solid')
            else:
                cell.fill = PatternFill(start_color=COLORS['row_green'], end_color=COLORS['row_green'], fill_type='solid')
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        
        current_row += 1
        count_table2 += 1
        ordem_counter += 1

def create_access_control_sheet(wb, df, vlan_id=50):
    """Cria planilha de Controle de Acesso com layout similar"""
    ws = wb.create_sheet('IP - CONTROLE DE ACESSO')
    
    # Configurar largura das colunas
    column_widths = [8, 12, 15, 15, 15, 20, 25, 40, 15, 18, 15, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 60
    
    # Título
    ws.merge_cells('A1:N1')
    title = ws['A1']
    title.value = 'LISTA DE EQUIPAMENTOS - CONDOMINIO CALABASAS - CONTROLE DE ACESSO'
    title.font = Font(bold=True, size=16, color=COLORS['text_dark'])
    title.fill = PatternFill(start_color=COLORS['title_bg'], end_color=COLORS['title_bg'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar logos
    try:
        logo_paths = ['logo_hi.png', 'logo_hi.jpg', 'logo_HI.png', 'logo_HI.jpg']
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                img_hi = Image(logo_path)
                img_hi.width = 100
                img_hi.height = 60
                ws.add_image(img_hi, 'A1')
                break
        
        calabasas_paths = ['logo_calabasas.png', 'logo_calabasas.jpg', 'Calabasas.png', 'Calabasas.jpg']
        for logo_path in calabasas_paths:
            if os.path.exists(logo_path):
                img_cal = Image(logo_path)
                img_cal.width = 150
                img_cal.height = 60
                ws.add_image(img_cal, 'M1')
                break
    except:
        pass
    
    # Encontrar cabeçalho
    header_row = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x) for x in row.values if pd.notna(x)]).upper()
        if 'ORDEM' in row_str and 'IP' in row_str:
            header_row = idx
            break
    if header_row is None:
        header_row = 4
    
    # Identificar colunas (similar ao CFTV)
    ordem_col = None
    tag_rack_col = None
    nomenclatura_col = None
    modelo_col = None
    localizacao_col = None
    ip_col = None
    usuario_op_col = None
    senha_op_col = None
    usuario_hi_col = None
    
    for col_idx in range(len(df.columns)):
        col_header = str(df.iloc[header_row, col_idx]).upper()
        if 'ORDEM' in col_header and ordem_col is None:
            ordem_col = col_idx
        if 'TAG' in col_header:
            tag_rack_col = col_idx
        if 'NOMENCLATURA' in col_header:
            nomenclatura_col = col_idx
        if 'MODELO' in col_header:
            modelo_col = col_idx
        if 'LOCALIZA' in col_header:
            localizacao_col = col_idx
        if 'IP' in col_header and ip_col is None and 'EQUIPAMENTO' not in col_header:
            ip_col = col_idx
        if 'USUARIO OPERADOR' in col_header or 'USUÁRIO OPERADOR' in col_header:
            usuario_op_col = col_idx
        if 'SENHA OPERADOR' in col_header:
            senha_op_col = col_idx
        if 'USUARIO HI' in col_header or 'USUÁRIO HI' in col_header:
            usuario_hi_col = col_idx
    
    def is_valid_data_row(row_data):
        if all(pd.isna(x) or str(x).strip() == '' for x in row_data):
            return False
        row_str = ' '.join([str(x) for x in row_data if pd.notna(x)]).upper()
        header_keywords = ['ORDEM', 'TAG', 'IP', 'MODELO', 'LOCALIZA', 'USUARIO', 'SENHA', 'NOMENCLATURA']
        keyword_count = sum(1 for keyword in header_keywords if keyword in row_str)
        if keyword_count >= 3:
            if ordem_col is not None and ordem_col < len(row_data):
                ordem_val = row_data[ordem_col]
                ordem_str = str(ordem_val).strip() if pd.notna(ordem_val) else ''
                if not ordem_str.isdigit():
                    return False
        if ip_col is not None and ip_col < len(row_data):
            ip_val = row_data[ip_col]
            ip_extracted = extract_ip(ip_val)
            if ip_extracted:
                return True
        if ordem_col is not None and ordem_col < len(row_data):
            ordem_val = row_data[ordem_col]
            ordem_str = str(ordem_val).strip() if pd.notna(ordem_val) else ''
            if ordem_str.isdigit():
                return True
        return False
    
    # Cabeçalho (verde para controle de acesso)
    current_row = 3
    headers = ['ORDEM', 'TAG DO RACK', 'NOMENCLATURA\nTAG DA CÂMERA', 'MODELO', 
              'LOCALIZAÇÃO CÂMERA', 'VLAN ID', 'VLAN Nome', 'IP ANTIGO', 'IP NOVO',
              'STATUS', 'USUARIO OPERADOR', 'SENHA OPERADOR', 'USUARIO HI', 'SENHA HI']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['header_green'], end_color=COLORS['header_green'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[current_row].height = 50
    current_row += 1
    
    # Dados
    count_rows = 0
    for idx in range(header_row + 1, len(df)):
        row_data = df.iloc[idx].values
        
        if not is_valid_data_row(row_data):
            continue
        
        ordem_val = row_data[ordem_col] if ordem_col is not None else None
        ip_val = row_data[ip_col] if ip_col is not None else None
        
        if pd.isna(ordem_val) or not str(ordem_val).strip().isdigit():
            continue
        if not extract_ip(ip_val):
            continue
        
        ordem = int(ordem_val) if pd.notna(ordem_val) else None
        tag_rack = row_data[tag_rack_col] if tag_rack_col is not None else None
        nomenclatura = row_data[nomenclatura_col] if nomenclatura_col is not None else None
        modelo = row_data[modelo_col] if modelo_col is not None else None
        localizacao = row_data[localizacao_col] if localizacao_col is not None else None
        ip_antigo = extract_ip(ip_val)
        
        vlan_id = get_vlan_for_equipment(tag_rack, modelo, localizacao, ip_antigo, 'ACCESS')
        vlan_name = 'Access Control' if vlan_id == 50 else 'CFTV' if vlan_id == 40 else 'Management'
        
        ip_novo = calculate_new_ip(ip_antigo, vlan_id, tag_rack)
        status = 'Migrado' if ip_novo else 'Pendente'
        
        usuario_op = row_data[usuario_op_col] if usuario_op_col is not None and pd.notna(row_data[usuario_op_col]) else 'operador'
        senha_op = row_data[senha_op_col] if senha_op_col is not None and pd.notna(row_data[senha_op_col]) else SENHA_OPERADOR
        usuario_hi = row_data[usuario_hi_col] if usuario_hi_col is not None and pd.notna(row_data[usuario_hi_col]) else 'admin'
        senha_hi = SENHA_HI
        
        row_values = [
            ordem, tag_rack, nomenclatura, modelo, localizacao,
            vlan_id, vlan_name, ip_antigo, ip_novo, status,
            usuario_op, senha_op, usuario_hi, senha_hi
        ]
        
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if count_rows % 2 == 0:
                cell.fill = PatternFill(start_color=COLORS['row_white'], end_color=COLORS['row_white'], fill_type='solid')
            else:
                cell.fill = PatternFill(start_color=COLORS['row_green'], end_color=COLORS['row_green'], fill_type='solid')
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        
        current_row += 1
        count_rows += 1

def main():
    print("Gerando Projeto Lógico de Rede - Calabasas...")
    
    # Resetar contadores
    global ip_counters
    ip_counters = {
        10: {'switches': 2, 'nvrs': 51, 'others': 100},
        40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
        50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
    }
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    excel_file = 'Cópia de IP Senhas Condominio Calabasas.xlsx'
    
    try:
        xls = pd.ExcelFile(excel_file)
        
        if 'IP - CFTV' in xls.sheet_names:
            df_cftv = pd.read_excel(excel_file, sheet_name='IP - CFTV', header=None)
            create_cftv_sheet(wb, df_cftv, 40)
        
        if 'IP - CONTROLE DE ACESSO' in xls.sheet_names:
            df_access = pd.read_excel(excel_file, sheet_name='IP - CONTROLE DE ACESSO', header=None)
            create_access_control_sheet(wb, df_access, 50)
        
        output_file = 'Projeto_Logico_Rede_Calabasas.xlsx'
        wb.save(output_file)
        print(f"Planilha gerada com sucesso: {output_file}")
        print(f"Senha HI atualizada para: {SENHA_HI}")
        
    except PermissionError as e:
        print(f"Erro: Arquivo '{excel_file}' está aberto. Feche o arquivo e tente novamente.")
    except Exception as e:
        print(f"Erro ao gerar planilha: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

