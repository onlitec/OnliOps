import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Cores profissionais
COLORS = {
    'header': '366092',
    'header_text': 'FFFFFF',
    'vlan_10': 'E2EFDA',
    'vlan_20': 'DEEBF7',
    'vlan_30': 'FFF2CC',
    'vlan_40': 'FCE4D6',
    'vlan_50': 'E1D5E7',
    'vlan_60': 'F4F4F4',
    'vlan_100': 'D9D9D9',
    'info': 'FFF9E6',
    'border': '000000',
    'warning': 'FFE699',
    'success': 'C6EFCE',
    'white': 'FFFFFF',
    'text_dark': '000000',
}

VLANS = {
    10: {'name': 'Management', 'subnet': '10.10.10.0/24', 'gateway': '10.10.10.1', 'color': COLORS['vlan_10'], 'desc': 'Gerenciamento de Infraestrutura'},
    20: {'name': 'Data', 'subnet': '10.10.20.0/24', 'gateway': '10.10.20.1', 'color': COLORS['vlan_20'], 'desc': 'Rede Corporativa'},
    30: {'name': 'Voice', 'subnet': '10.10.30.0/24', 'gateway': '10.10.30.1', 'color': COLORS['vlan_30'], 'desc': 'Telefonia IP'},
    40: {'name': 'CFTV', 'subnet': '10.10.40.0/24', 'gateway': '10.10.40.1', 'color': COLORS['vlan_40'], 'desc': 'Sistema de CFTV'},
    50: {'name': 'Access Control', 'subnet': '10.10.50.0/24', 'gateway': '10.10.50.1', 'color': COLORS['vlan_50'], 'desc': 'Controle de Acesso'},
    60: {'name': 'IoT', 'subnet': '10.10.60.0/24', 'gateway': '10.10.60.1', 'color': COLORS['vlan_60'], 'desc': 'Dispositivos IoT'},
    100: {'name': 'Guest', 'subnet': '10.10.100.0/24', 'gateway': '10.10.100.1', 'color': COLORS['vlan_100'], 'desc': 'WiFi Visitantes'}
}

# Contadores de IP
ip_counters = {
    10: {'switches': 2, 'nvrs': 51, 'others': 100},
    40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
    50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
}

def extract_ip(ip_str):
    if pd.isna(ip_str):
        return None
    ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    match = re.search(ip_pattern, str(ip_str))
    return match.group(0) if match else None

def get_vlan_for_equipment(tag, modelo, localizacao, ip_atual):
    if pd.isna(tag):
        tag = ''
    if pd.isna(modelo):
        modelo = ''
    if pd.isna(localizacao):
        localizacao = ''
    
    tag_str = str(tag).upper()
    modelo_str = str(modelo).upper()
    localizacao_str = str(localizacao).upper()
    
    if any(x in tag_str for x in ['SW-', 'RK-', 'ROTEADOR']):
        return 10
    if 'NVR' in tag_str or 'NVR' in modelo_str:
        return 10
    if 'CA-' in tag_str and any(x in localizacao_str for x in ['PERIMETRO', 'ESTACIONAMENTO', 'TORRE', 'AVENIDA']):
        return 40
    if 'EL-' in tag_str or 'ELEVADOR' in localizacao_str:
        return 50
    if 'LPR' in tag_str or 'LPR' in localizacao_str:
        return 50
    if 'LF-' in tag_str or 'LEITOR' in modelo_str:
        return 50
    if 'CE-' in tag_str or 'CONTROLADOR' in modelo_str:
        return 50
    if 'PORTARIA' in localizacao_str or 'PORTÃO' in localizacao_str:
        return 50
    if 'CM-' in tag_str or 'CONVERSOR' in modelo_str:
        return 40
    if ip_atual and ip_atual.startswith('10.10.0.'):
        return 40
    if ip_atual and ip_atual.startswith('10.10.1.'):
        return 50
    return 40

def calculate_new_ip(old_ip, vlan_id, equipment_type=None):
    global ip_counters
    if not old_ip:
        return None
    try:
        last_octet = int(old_ip.split('.')[-1])
        if vlan_id == 10:
            if 'SW' in str(equipment_type).upper() or 'ROTEADOR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['switches']
                ip_counters[10]['switches'] = min(ip_counters[10]['switches'] + 1, 50)
            elif 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['nvrs']
                ip_counters[10]['nvrs'] = min(ip_counters[10]['nvrs'] + 1, 100)
            else:
                new_octet = ip_counters[10]['others']
                ip_counters[10]['others'] = min(ip_counters[10]['others'] + 1, 254)
            return f'10.10.10.{new_octet}'
        elif vlan_id == 40:
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['nvrs']
                ip_counters[40]['nvrs'] = min(ip_counters[40]['nvrs'] + 1, 10)
            elif 'CM-' in str(equipment_type).upper() or 'CONVERSOR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['converters']
                ip_counters[40]['converters'] = min(ip_counters[40]['converters'] + 1, 240)
            else:
                new_octet = ip_counters[40]['cameras']
                ip_counters[40]['cameras'] = min(ip_counters[40]['cameras'] + 1, 200)
            return f'10.10.40.{new_octet}'
        elif vlan_id == 50:
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['nvrs']
                ip_counters[50]['nvrs'] = min(ip_counters[50]['nvrs'] + 1, 10)
            elif 'EL-' in str(equipment_type).upper() or 'ELEVADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['cameras_elevator']
                ip_counters[50]['cameras_elevator'] = min(ip_counters[50]['cameras_elevator'] + 1, 50)
            elif 'LF-' in str(equipment_type).upper() or 'LEITOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['readers']
                ip_counters[50]['readers'] = min(ip_counters[50]['readers'] + 1, 200)
            elif 'CE-' in str(equipment_type).upper() or 'CONTROLADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['controllers']
                ip_counters[50]['controllers'] = min(ip_counters[50]['controllers'] + 1, 240)
            else:
                new_octet = ip_counters[50]['cameras_portaria']
                ip_counters[50]['cameras_portaria'] = min(ip_counters[50]['cameras_portaria'] + 1, 100)
            return f'10.10.50.{new_octet}'
        else:
            base = vlan_id * 10 if vlan_id < 100 else 100
            return f'10.10.{base}.{last_octet}'
    except:
        return None

def apply_border(cell):
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )
    cell.border = thin_border

def create_summary_sheet(wb):
    ws = wb.create_sheet('RESUMO VLANs', 0)
    
    # Título
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = 'ARQUITETURA DE REDE COM SEGMENTAÇÃO POR VLANS - CONDOMÍNIO CALABASAS'
    title.font = Font(bold=True, size=16, color=COLORS['header_text'])
    title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title)
    ws.row_dimensions[1].height = 35
    
    # Info
    ws.merge_cells('A2:H2')
    info = ws['A2']
    info.value = f'Documento gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Servidor HikCentral: 10.10.20.10 (VLAN 20 - Data)'
    info.font = Font(size=10, italic=True, bold=True)
    info.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    info.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(info)
    ws.row_dimensions[2].height = 25
    
    # Cabeçalhos
    headers = ['VLAN ID', 'Nome', 'Sub-rede', 'Gateway', 'Máscara', 'Descrição', 'Faixa de IPs', 'Cor']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, size=11, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 45
    
    # Dados VLANs
    vlan_data = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', '255.255.255.0', 'Gerenciamento de infraestrutura', '10.10.10.2-254', 'Verde claro'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', '255.255.255.0', 'Rede corporativa (HikCentral)', '10.10.20.2-254', 'Azul claro'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', '255.255.255.0', 'Telefonia IP', '10.10.30.2-254', 'Amarelo claro'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', '255.255.255.0', 'Sistema de CFTV', '10.10.40.2-254', 'Laranja claro'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', '255.255.255.0', 'Controle de acesso', '10.10.50.2-254', 'Roxo claro'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', '255.255.255.0', 'Dispositivos IoT', '10.10.60.2-254', 'Cinza claro'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', '255.255.255.0', 'WiFi visitantes', '10.10.100.2-254', 'Cinza']
    ]
    
    for row_idx, row_data in enumerate(vlan_data, start=4):
        vlan_id = row_data[0]
        vlan_info = VLANS[vlan_id]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=vlan_info['color'], end_color=vlan_info['color'], fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 28
    
    # Legenda
    start_row = len(vlan_data) + 6
    ws.merge_cells(f'A{start_row}:H{start_row}')
    legend_title = ws.cell(row=start_row, column=1)
    legend_title.value = 'LEGENDA E INFORMAÇÕES IMPORTANTES'
    legend_title.font = Font(bold=True, size=12, color=COLORS['header_text'])
    legend_title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    legend_title.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(legend_title)
    ws.row_dimensions[start_row].height = 28
    
    legend_items = [
        ['SERVIDOR HIKCENTRAL:', 'IP: 10.10.20.10 na VLAN 20 (Data). Gerencia CFTV (VLAN 40) e Controle de Acesso (VLAN 50)'],
        ['CORES:', 'Cada VLAN possui uma cor específica para facilitar identificação visual nas planilhas'],
        ['IPs ESTÁTICOS:', 'Todos os equipamentos de CFTV e Controle de Acesso devem usar IPs estáticos'],
        ['MIGRAÇÃO:', 'Consulte a planilha "MAPEAMENTO MIGRACAO" para ver IP antigo → IP novo'],
        ['CREDENCIAIS:', 'Usuário Operador: operador | Senha: cc2025 | Usuário HI: admin | Senha: Hical@20#25'],
        ['ATENÇÃO:', 'Alterar senhas padrão após migração para maior segurança']
    ]
    
    for idx, (label, desc) in enumerate(legend_items, start=start_row + 1):
        ws.merge_cells(f'A{idx}:B{idx}')
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        apply_border(label_cell)
        
        ws.merge_cells(f'C{idx}:H{idx}')
        desc_cell = ws.cell(row=idx, column=3, value=desc)
        desc_cell.font = Font(size=10)
        desc_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        apply_border(desc_cell)
        ws.row_dimensions[idx].height = 25
    
    column_widths = [12, 18, 18, 15, 15, 35, 25, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def process_equipment_sheet_detailed(wb, sheet_name, df, vlan_id, vlan_name):
    """Processa planilha de equipamentos com layout similar à imagem"""
    ws = wb.create_sheet(f'{sheet_name} (VLAN {vlan_id})')
    
    # Configurar largura das colunas
    column_widths = [8, 12, 20, 25, 40, 15, 18, 15, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Altura da linha do título
    ws.row_dimensions[1].height = 60
    
    # Título principal (linha 1) - estilo da imagem
    ws.merge_cells('A1:K1')
    title = ws['A1']
    title.value = 'LISTA DE EQUIPAMENTOS - CONDOMINIO CALABASAS'
    title.font = Font(bold=True, size=16, color=COLORS['text_dark'])
    title.fill = PatternFill(start_color=COLORS['white'], end_color=COLORS['white'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar logos (se existirem)
    try:
        from openpyxl.drawing.image import Image
        # Logo HI (lado esquerdo)
        logo_paths = ['logo_hi.png', 'logo_hi.jpg', 'logo_HI.png', 'logo_HI.jpg']
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                img_hi = Image(logo_path)
                img_hi.width = 100
                img_hi.height = 60
                ws.add_image(img_hi, 'A1')
                break
        
        # Logo Calabasas (lado direito)
        calabasas_paths = ['logo_calabasas.png', 'logo_calabasas.jpg', 'Calabasas.png', 'Calabasas.jpg']
        for logo_path in calabasas_paths:
            if os.path.exists(logo_path):
                img_cal = Image(logo_path)
                img_cal.width = 150
                img_cal.height = 60
                ws.add_image(img_cal, 'J1')
                break
    except Exception as e:
        print(f"Nota: Logos não encontrados ou erro ao adicionar: {e}")
        print("Para adicionar logos, coloque os arquivos logo_hi.png e logo_calabasas.png na pasta do projeto")
    
    # Espaço após título
    current_row = 3
    
    # Encontrar cabeçalho
    header_row = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x) for x in row.values if pd.notna(x)]).upper()
        if 'ORDEM' in row_str and 'IP' in row_str:
            header_row = idx
            break
    if header_row is None:
        header_row = 4
    
    # Cabeçalhos - mapear corretamente as colunas (ignorar coluna 0 que é vazia)
    original_headers = []
    original_col_indices = []
    for col_idx in range(len(df.columns)):
        header_val = df.iloc[header_row, col_idx]
        if pd.notna(header_val) and str(header_val).strip():
            original_headers.append(str(header_val))
            original_col_indices.append(col_idx)
    
    # Cabeçalhos da primeira tabela (vermelho - estilo imagem)
    headers_table1 = ['ORDEM', 'TAG DO RACK', 'NOMENCLATURA\nTAG DA CÂMERA', 'MODELO', 
                      'LOCALIZAÇÃO CÂMERA', 'IP', 'USUARIO OPERADOR', 'SENHA OPERADOR', 
                      'USUARIO HI', 'SENHA HI']
    
    for col_idx, h in enumerate(headers_table1, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[current_row].height = 40
    current_row += 1
    
    # Identificar colunas
    ip_col_idx = None
    tag_col_idx = None
    nomenclatura_col_idx = None
    modelo_col_idx = None
    localizacao_col_idx = None
    
    for col_idx in range(len(df.columns)):
        col_header = str(df.iloc[header_row, col_idx]).upper()
        if 'IP' in col_header and ip_col_idx is None and 'EQUIPAMENTO' not in col_header:
            ip_col_idx = col_idx
        if 'TAG' in col_header and ('RACK' in col_header or 'CAIXA' in col_header):
            tag_col_idx = col_idx
        if 'NOMENCLATURA' in col_header:
            nomenclatura_col_idx = col_idx
        if 'MODELO' in col_header:
            modelo_col_idx = col_idx
        if 'LOCALIZA' in col_header:
            localizacao_col_idx = col_idx
    
    # Identificar coluna de ordem
    ordem_col_idx = None
    for col_idx in range(len(df.columns)):
        col_header = str(df.iloc[header_row, col_idx]).upper()
        if 'ORDEM' in col_header and ordem_col_idx is None:
            ordem_col_idx = col_idx
            break
    
    # Função para validar se linha é válida (não é cabeçalho repetido)
    def is_valid_data_row(row_data):
        """Verifica se uma linha contém dados válidos e não é cabeçalho repetido"""
        # Verificar se linha está completamente vazia
        if all(pd.isna(x) or str(x).strip() == '' for x in row_data):
            return False
        
        # Verificar se é cabeçalho repetido
        row_str = ' '.join([str(x) for x in row_data if pd.notna(x)]).upper()
        header_keywords = ['ORDEM', 'TAG', 'IP', 'MODELO', 'LOCALIZA', 'USUARIO', 'SENHA', 'NOMENCLATURA']
        
        # Se a linha contém muitas palavras de cabeçalho, provavelmente é cabeçalho repetido
        keyword_count = sum(1 for keyword in header_keywords if keyword in row_str)
        if keyword_count >= 3:
            # Verificar se não tem número de ordem (cabeçalhos não têm números)
            if ordem_col_idx is not None and ordem_col_idx < len(row_data):
                ordem_val = row_data[ordem_col_idx]
                ordem_str = str(ordem_val).strip() if pd.notna(ordem_val) else ''
                if not ordem_str.isdigit():
                    return False
        
        # Verificar se tem IP válido OU ordem numérica válida
        has_valid_ip = False
        if ip_col_idx is not None and ip_col_idx < len(row_data):
            ip_val = row_data[ip_col_idx]
            ip_extracted = extract_ip(ip_val)
            if ip_extracted:
                has_valid_ip = True
        
        has_valid_ordem = False
        if ordem_col_idx is not None and ordem_col_idx < len(row_data):
            ordem_val = row_data[ordem_col_idx]
            ordem_str = str(ordem_val).strip() if pd.notna(ordem_val) else ''
            if ordem_str.isdigit():
                has_valid_ordem = True
        
        # Linha é válida se tem IP válido OU ordem numérica válida
        return has_valid_ip or has_valid_ordem
    
    # Processar dados - primeira tabela (primeiras 20 linhas)
    count_rows = 0
    for idx in range(header_row + 1, len(df)):
        if count_rows >= 20:
            break
        
        row_data = df.iloc[idx].values
        
        # Validar se linha é válida
        if not is_valid_data_row(row_data):
            continue
        
        ip_antigo = extract_ip(row_data[ip_col_idx]) if ip_col_idx is not None and ip_col_idx < len(row_data) else None
        
        # Se não tem IP antigo, pular (não é um equipamento válido)
        if not ip_antigo:
            continue
        
        # Extrair dados na ordem correta
        ordem_val = row_data[ordem_col_idx] if ordem_col_idx is not None and ordem_col_idx < len(row_data) else None
        tag_val = row_data[tag_col_idx] if tag_col_idx is not None and tag_col_idx < len(row_data) else None
        nomenclatura_val = row_data[nomenclatura_col_idx] if nomenclatura_col_idx is not None and nomenclatura_col_idx < len(row_data) else None
        modelo_val = row_data[modelo_col_idx] if modelo_col_idx is not None and modelo_col_idx < len(row_data) else None
        localizacao_val = row_data[localizacao_col_idx] if localizacao_col_idx is not None and localizacao_col_idx < len(row_data) else None
        
        # Buscar credenciais
        usuario_op_val = None
        senha_op_val = None
        usuario_hi_val = None
        senha_hi_val = None
        
        for col_idx in original_col_indices:
            if col_idx < len(row_data):
                col_header = str(df.iloc[header_row, col_idx]).upper()
                if 'USUARIO OPERADOR' in col_header or 'USUÁRIO OPERADOR' in col_header:
                    usuario_op_val = row_data[col_idx] if pd.notna(row_data[col_idx]) else 'operador'
                if 'SENHA OPERADOR' in col_header:
                    senha_op_val = row_data[col_idx] if pd.notna(row_data[col_idx]) else 'cc2025'
                if 'USUARIO HI' in col_header or 'USUÁRIO HI' in col_header:
                    usuario_hi_val = row_data[col_idx] if pd.notna(row_data[col_idx]) else 'admin'
                if 'SENHA HI' in col_header:
                    senha_hi_val = 'Hical@20#25'  # Sempre usar senha atualizada
        
        # Construir linha com dados na ordem da imagem
        row_values = [
            ordem_val,
            tag_val,
            nomenclatura_val,
            modelo_val,
            localizacao_val,
            ip_antigo,
            usuario_op_val if usuario_op_val else 'operador',
            senha_op_val if senha_op_val else 'cc2025',
            usuario_hi_val if usuario_hi_val else 'admin',
            'Hical@20#25'  # Sempre usar senha atualizada
        ]
        
        # Adicionar linha
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            # Linhas alternadas (branco e verde claro) - estilo da imagem
            row_index = current_row - 4  # Ajustar índice (cabeçalho na linha 3)
            if row_index % 2 == 0:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        
        current_row += 1
        count_rows += 1
    
    # Segunda tabela (se houver mais dados) - cabeçalho verde
    if count_rows < len([r for r in range(header_row + 1, len(df)) if is_valid_data_row(df.iloc[r].values)]):
        current_row += 2  # Espaço entre tabelas
        
        headers_table2 = ['ORDEM', 'TAG DA CAIXA', 'NOMENCLATURA\nTAG DA CÂMERA', 'MODELO', 
                          'LOCALIZAÇÃO CÂMERA', 'IP', 'USUARIO OPERADOR', 'SENHA OPERADOR', 
                          'USUARIO HI', 'SENHA HI']
        
        for col_idx, h in enumerate(headers_table2, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = h
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[current_row].height = 40
        current_row += 1
        
        # Processar dados restantes
        ordem_counter = 1
        for idx in range(header_row + 1, len(df)):
            row_data = df.iloc[idx].values
            
            # Pular primeiras 20 linhas já processadas
            ordem_val = row_data[ordem_col_idx] if ordem_col_idx is not None and ordem_col_idx < len(row_data) else None
            if pd.notna(ordem_val) and str(ordem_val).strip().isdigit():
                if int(ordem_val) <= 20:
                    continue
            
            if not is_valid_data_row(row_data):
                continue
            
            ip_antigo = extract_ip(row_data[ip_col_idx]) if ip_col_idx is not None and ip_col_idx < len(row_data) else None
            if not ip_antigo:
                continue
            
            # Extrair dados
            tag_val = row_data[tag_col_idx] if tag_col_idx is not None and tag_col_idx < len(row_data) else None
            nomenclatura_val = row_data[nomenclatura_col_idx] if nomenclatura_col_idx is not None and nomenclatura_col_idx < len(row_data) else None
            modelo_val = row_data[modelo_col_idx] if modelo_col_idx is not None and modelo_col_idx < len(row_data) else None
            localizacao_val = row_data[localizacao_col_idx] if localizacao_col_idx is not None and localizacao_col_idx < len(row_data) else None
            
            # Buscar credenciais
            usuario_op_val = 'operador'
            senha_op_val = 'cc2025'
            usuario_hi_val = 'admin'
            
            for col_idx in original_col_indices:
                if col_idx < len(row_data):
                    col_header = str(df.iloc[header_row, col_idx]).upper()
                    if 'USUARIO OPERADOR' in col_header or 'USUÁRIO OPERADOR' in col_header:
                        usuario_op_val = row_data[col_idx] if pd.notna(row_data[col_idx]) else 'operador'
                    if 'SENHA OPERADOR' in col_header:
                        senha_op_val = row_data[col_idx] if pd.notna(row_data[col_idx]) else 'cc2025'
                    if 'USUARIO HI' in col_header or 'USUÁRIO HI' in col_header:
                        usuario_hi_val = row_data[col_idx] if pd.notna(row_data[col_idx]) else 'admin'
            
            row_values = [
                ordem_counter,
                tag_val,
                nomenclatura_val,
                modelo_val,
                localizacao_val,
                ip_antigo,
                usuario_op_val,
                senha_op_val,
                usuario_hi_val,
                'Hical@20#25'  # Sempre usar senha atualizada
            ]
            
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                row_index = current_row - (count_rows + 5)  # Ajustar índice
                if row_index % 2 == 0:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                apply_border(cell)
            
            current_row += 1
            ordem_counter += 1

def create_management_sheet(wb):
    ws = wb.create_sheet('GERENCIAMENTO (VLAN 10)')
    
    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = 'EQUIPAMENTOS DE GERENCIAMENTO - VLAN 10 (Management)'
    title.font = Font(bold=True, size=14, color=COLORS['header_text'])
    title.fill = PatternFill(start_color=VLANS[10]['color'], end_color=VLANS[10]['color'], fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title)
    ws.row_dimensions[1].height = 35
    
    headers = ['VLAN ID', 'VLAN Nome', 'IP', 'TAG', 'Equipamento', 'Modelo', 'Localização', 'Função']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, size=11, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_border(cell)
    ws.row_dimensions[2].height = 40
    
    switches = [
        [10, 'Management', '10.10.10.2', 'SW-00-RK-00', 'Switch Principal Rack', 'DS-3E1526P-EI', 'Rack Principal', 'Switch Core'],
        [10, 'Management', '10.10.10.3', 'SW-01-TR-01', 'Switch Torre 1', 'DS-3E1318P-EI/M', 'Torre 1', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.4', 'SW-02-TR-02', 'Switch Torre 2', 'DS-3E1526P-EI', 'Torre 2', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.5', 'SW-03-TR-03', 'Switch Torre 3', 'DS-3E1526P-EI', 'Torre 3', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.6', 'SW-04-TR-04', 'Switch Torre 4', 'DS-3E1526P-EI', 'Torre 4', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.7', 'SW-01-CP-01', 'Switch Perimetral 1', 'DS-3E1309P-EI', 'Caixa Perimetral 1', 'Switch Acesso'],
        [10, 'Management', '10.10.10.8', 'SW-02-CP-02', 'Switch Perimetral 2', 'DS-3E1309P-EI', 'Caixa Perimetral 2', 'Switch Acesso'],
        [10, 'Management', '10.10.10.9', 'SW-03-CP-03', 'Switch Perimetral 3', 'DS-3E1309P-EI', 'Caixa Perimetral 3', 'Switch Acesso'],
        [10, 'Management', '10.10.10.10', 'SW-04-CP-04', 'Switch Perimetral 4', 'DS-3E1309P-EI', 'Caixa Perimetral 4', 'Switch Acesso'],
        [10, 'Management', '10.10.10.1', 'ROTEADOR-01', 'Roteador/Gateway', 'MIKROTIK RB750Gr3', 'Rack Principal', 'Gateway'],
    ]
    
    for row_data in switches:
        ws.append(row_data)
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=VLANS[10]['color'], end_color=VLANS[10]['color'], fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_num].height = 25
    
    column_widths = [12, 18, 15, 15, 25, 20, 25, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def main():
    print("Gerando planilha Excel completa e organizada...")
    
    global ip_counters
    ip_counters = {
        10: {'switches': 2, 'nvrs': 51, 'others': 100},
        40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
        50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
    }
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    excel_file = 'Cópia de IP Senhas Condominio Calabasas.xlsx'
    xls = pd.ExcelFile(excel_file)
    
    create_summary_sheet(wb)
    
    if 'IP - CFTV' in xls.sheet_names:
        df_cftv = pd.read_excel(excel_file, sheet_name='IP - CFTV', header=None)
        process_equipment_sheet_detailed(wb, 'IP - CFTV', df_cftv, 40, 'CFTV')
    
    if 'IP - CONTROLE DE ACESSO' in xls.sheet_names:
        df_access = pd.read_excel(excel_file, sheet_name='IP - CONTROLE DE ACESSO', header=None)
        process_equipment_sheet_detailed(wb, 'IP - CONTROLE ACESSO', df_access, 50, 'Access Control')
    
    create_management_sheet(wb)
    
    output_file = 'Rede_VLANs_Condominio_Calabasas.xlsx'
    try:
        wb.save(output_file)
        print(f"Planilha gerada com sucesso: {output_file}")
    except PermissionError:
        output_file_temp = 'Rede_VLANs_Condominio_Calabasas_TEMP.xlsx'
        wb.save(output_file_temp)
        print(f"Arquivo está aberto. Planilha salva como: {output_file_temp}")
        print(f"Feche o arquivo {output_file} e renomeie {output_file_temp} para {output_file}")

if __name__ == '__main__':
    main()

