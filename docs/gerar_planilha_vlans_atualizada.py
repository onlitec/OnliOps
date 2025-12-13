import pandas as pd
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import hashlib

# Cores baseadas em planilhas profissionais (tons suaves)
COLORS = {
    'header': '2F4F4F',  # Verde-azulado escuro para cabeçalhos
    'header_text': 'FFFFFF',  # Branco para texto de cabeçalho
    'vlan_10': 'E2EFDA',  # Verde muito claro - Management
    'vlan_20': 'DEEBF7',  # Azul muito claro - Data
    'vlan_30': 'FFF2CC',  # Amarelo muito claro - Voice
    'vlan_40': 'FCE4D6',  # Laranja muito claro - CFTV
    'vlan_50': 'E1D5E7',  # Roxo muito claro - Access Control
    'vlan_60': 'F4F4F4',  # Cinza muito claro - IoT
    'vlan_100': 'D9D9D9',  # Cinza claro - Guest
    'info': 'FFF9E6',  # Bege claro para informações
    'border': '000000',  # Preto para bordas
    'text_dark': '000000',  # Preto para texto
    'text_light': 'FFFFFF',  # Branco para texto claro
    'warning': 'FFE699',  # Amarelo para avisos
    'success': 'C6EFCE',  # Verde claro para sucesso
    'error': 'FFC7CE',    # Vermelho claro para erros
    'security': 'FF9999', # Vermelho para alertas de segurança
}

# Mapeamento de VLANs atualizado
VLANS = {
    10: {'name': 'Management', 'subnet': '10.10.10.0/24', 'gateway': '10.10.10.1', 'color': COLORS['vlan_10'], 'desc': 'Gerenciamento de Infraestrutura'},
    20: {'name': 'Data', 'subnet': '10.10.20.0/24', 'gateway': '10.10.20.1', 'color': COLORS['vlan_20'], 'desc': 'Rede Corporativa'},
    30: {'name': 'Voice', 'subnet': '10.10.30.0/24', 'gateway': '10.10.30.1', 'color': COLORS['vlan_30'], 'desc': 'Telefonia IP'},
    40: {'name': 'CFTV', 'subnet': '10.10.40.0/24', 'gateway': '10.10.40.1', 'color': COLORS['vlan_40'], 'desc': 'Sistema de CFTV'},
    50: {'name': 'Access Control', 'subnet': '10.10.50.0/24', 'gateway': '10.10.50.1', 'color': COLORS['vlan_50'], 'desc': 'Controle de Acesso'},
    60: {'name': 'IoT', 'subnet': '10.10.60.0/24', 'gateway': '10.10.60.1', 'color': COLORS['vlan_60'], 'desc': 'Dispositivos IoT'},
    100: {'name': 'Guest', 'subnet': '10.10.100.0/24', 'gateway': '10.10.100.1', 'color': COLORS['vlan_100'], 'desc': 'WiFi Visitantes'}
}

# Contadores para IPs sequenciais
ip_counters = {
    10: {'switches': 2, 'nvrs': 51, 'others': 100},
    40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
    50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
}

# Armazenamento de IPs para detecção de duplicatas
assigned_ips = set()

def extract_ip(ip_str):
    """Extrai IP de uma string"""
    if pd.isna(ip_str):
        return None
    ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    match = re.search(ip_pattern, str(ip_str))
    return match.group(0) if match else None

def is_valid_ip(ip):
    """Verifica se um IP é válido"""
    if not ip:
        return False
    parts = ip.split('.')
    if len(parts) != 4:
        return False
    for part in parts:
        try:
            num = int(part)
            if num < 0 or num > 255:
                return False
        except ValueError:
            return False
    return True

def get_vlan_for_equipment(tag, modelo, localizacao, ip_atual):
    """Determina a VLAN baseado no tipo de equipamento"""
    if pd.isna(tag):
        tag = ''
    if pd.isna(modelo):
        modelo = ''
    if pd.isna(localizacao):
        localizacao = ''
    
    tag_str = str(tag).upper()
    modelo_str = str(modelo).upper()
    localizacao_str = str(localizacao).upper()
    
    if any(x in tag_str for x in ['SW-', 'RK-', 'ROTEADOR']):
        return 10
    if 'NVR' in tag_str or 'NVR' in modelo_str:
        return 10
    if 'CA-' in tag_str and any(x in localizacao_str for x in ['PERIMETRO', 'ESTACIONAMENTO', 'TORRE', 'AVENIDA']):
        return 40
    if 'EL-' in tag_str or 'ELEVADOR' in localizacao_str:
        return 50
    if 'LPR' in tag_str or 'LPR' in localizacao_str:
        return 50
    if 'LF-' in tag_str or 'LEITOR' in modelo_str:
        return 50
    if 'CE-' in tag_str or 'CONTROLADOR' in modelo_str:
        return 50
    if 'PORTARIA' in localizacao_str or 'PORTÃO' in localizacao_str:
        return 50
    if 'CM-' in tag_str or 'CONVERSOR' in modelo_str:
        return 40
    if ip_atual and ip_atual.startswith('10.10.0.'):
        return 40
    if ip_atual and ip_atual.startswith('10.10.1.'):
        return 50
    return 40

def calculate_new_ip(old_ip, vlan_id, equipment_type=None):
    """Calcula novo IP baseado na VLAN e tipo de equipamento"""
    global ip_counters, assigned_ips
    
    if not old_ip:
        return None
    
    try:
        # Verificar se o IP já foi atribuído
        base_ip = f'10.10.{vlan_id}.'
        
        if vlan_id == 10:
            if 'SW' in str(equipment_type).upper() or 'ROTEADOR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['switches']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 50:
                    new_octet += 1
                ip_counters[10]['switches'] = min(new_octet + 1, 50)
            elif 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['nvrs']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 100:
                    new_octet += 1
                ip_counters[10]['nvrs'] = min(new_octet + 1, 100)
            else:
                new_octet = ip_counters[10]['others']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 254:
                    new_octet += 1
                ip_counters[10]['others'] = min(new_octet + 1, 254)
            new_ip = f'10.10.10.{new_octet}'
        
        elif vlan_id == 40:
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['nvrs']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 10:
                    new_octet += 1
                ip_counters[40]['nvrs'] = min(new_octet + 1, 10)
            elif 'CM-' in str(equipment_type).upper() or 'CONVERSOR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['converters']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 240:
                    new_octet += 1
                ip_counters[40]['converters'] = min(new_octet + 1, 240)
            else:
                new_octet = ip_counters[40]['cameras']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 200:
                    new_octet += 1
                ip_counters[40]['cameras'] = min(new_octet + 1, 200)
            new_ip = f'10.10.40.{new_octet}'
        
        elif vlan_id == 50:
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['nvrs']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 10:
                    new_octet += 1
                ip_counters[50]['nvrs'] = min(new_octet + 1, 10)
            elif 'EL-' in str(equipment_type).upper() or 'ELEVADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['cameras_elevator']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 50:
                    new_octet += 1
                ip_counters[50]['cameras_elevator'] = min(new_octet + 1, 50)
            elif 'LF-' in str(equipment_type).upper() or 'LEITOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['readers']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 200:
                    new_octet += 1
                ip_counters[50]['readers'] = min(new_octet + 1, 200)
            elif 'CE-' in str(equipment_type).upper() or 'CONTROLADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['controllers']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 240:
                    new_octet += 1
                ip_counters[50]['controllers'] = min(new_octet + 1, 240)
            else:
                new_octet = ip_counters[50]['cameras_portaria']
                while f'{base_ip}{new_octet}' in assigned_ips and new_octet <= 100:
                    new_octet += 1
                ip_counters[50]['cameras_portaria'] = min(new_octet + 1, 100)
            new_ip = f'10.10.50.{new_octet}'
        
        else:
            base = vlan_id * 10 if vlan_id < 100 else 100
            last_octet = int(old_ip.split('.')[-1])
            new_ip = f'10.10.{base}.{last_octet}'
        
        # Verificar se o IP é válido e não duplicado
        if is_valid_ip(new_ip) and new_ip not in assigned_ips:
            assigned_ips.add(new_ip)
            return new_ip
        else:
            # Tentar encontrar um IP disponível
            for i in range(2, 254):
                test_ip = f'{base_ip}{i}'
                if test_ip not in assigned_ips and is_valid_ip(test_ip):
                    assigned_ips.add(test_ip)
                    return test_ip
            return None
    
    except Exception as e:
        print(f"Erro ao calcular IP: {e}")
        return None

def apply_border(cell):
    """Aplica borda fina à célula"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )
    cell.border = thin_border

def create_summary_sheet(wb):
    """Cria planilha de resumo de VLANs com legendas"""
    ws = wb.create_sheet('Resumo Geral', 0)
    
    # Título principal
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'ARQUITETURA DE REDE COM SEGMENTAÇÃO POR VLANS - CONDOMÍNIO CALABASAS'
    title_cell.font = Font(bold=True, size=16, color=COLORS['header_text'])
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Informações gerais
    ws.merge_cells('A2:H2')
    info_cell = ws['A2']
    info_cell.value = f'Documento gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Servidor HikCentral: 10.10.20.10 (VLAN 20)'
    info_cell.font = Font(size=10, italic=True)
    info_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(info_cell)
    ws.row_dimensions[2].height = 20
    
    # Cabeçalho da tabela
    headers = ['VLAN ID', 'Nome', 'Sub-rede', 'Gateway', 'Máscara', 'Descrição', 'Faixa de IPs', 'Cor']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 40
    
    # Dados das VLANs
    vlan_data = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', '255.255.255.0', 'Gerenciamento de infraestrutura', '10.10.10.2-254', 'Verde claro'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', '255.255.255.0', 'Rede corporativa (HikCentral)', '10.10.20.2-254', 'Azul claro'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', '255.255.255.0', 'Telefonia IP', '10.10.30.2-254', 'Amarelo claro'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', '255.255.255.0', 'Sistema de CFTV', '10.10.40.2-254', 'Laranja claro'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', '255.255.255.0', 'Controle de acesso', '10.10.50.2-254', 'Roxo claro'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', '255.255.255.0', 'Dispositivos IoT', '10.10.60.2-254', 'Cinza claro'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', '255.255.255.0', 'WiFi visitantes', '10.10.100.2-254', 'Cinza']
    ]
    
    for row_idx, row_data in enumerate(vlan_data, start=4):
        vlan_id = row_data[0]
        vlan_info = VLANS[vlan_id]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=vlan_info['color'], end_color=vlan_info['color'], fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 25
    
    # Legenda e informações adicionais
    start_row = len(vlan_data) + 6
    ws.merge_cells(f'A{start_row}:H{start_row}')
    legend_title = ws.cell(row=start_row, column=1)
    legend_title.value = 'LEGENDA E INFORMAÇÕES IMPORTANTES'
    legend_title.font = Font(bold=True, size=12, color=COLORS['header_text'])
    legend_title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    legend_title.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(legend_title)
    ws.row_dimensions[start_row].height = 25
    
    legend_items = [
        ['SERVIDOR HIKCENTRAL:', 'IP: 10.10.20.10 na VLAN 20 (Data). Gerencia CFTV (VLAN 40) e Controle de Acesso (VLAN 50)'],
        ['CORES:', 'Cada VLAN possui uma cor específica para facilitar identificação visual nas planilhas'],
        ['IPs ESTÁTICOS:', 'Todos os equipamentos de CFTV e Controle de Acesso devem usar IPs estáticos'],
        ['MIGRAÇÃO:', 'Consulte a planilha "MAPEAMENTO MIGRACAO" para ver IP antigo → IP novo'],
        ['SEGURANÇA:', 'Credenciais padrão devem ser alteradas imediatamente após a migração'],
        ['BACKUP:', 'Manter cópias de segurança da configuração de todos os equipamentos']
    ]
    
    for idx, (label, desc) in enumerate(legend_items, start=start_row + 1):
        ws.merge_cells(f'A{idx}:B{idx}')
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        apply_border(label_cell)
        
        ws.merge_cells(f'C{idx}:H{idx}')
        desc_cell = ws.cell(row=idx, column=3, value=desc)
        desc_cell.font = Font(size=10)
        desc_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        apply_border(desc_cell)
        ws.row_dimensions[idx].height = 20
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 18, 15, 15, 35, 25, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def process_equipment_sheet(wb, sheet_name, df, vlan_id, vlan_name):
    """Processa planilha de equipamentos com formatação melhorada"""
    ws = wb.create_sheet(f'{sheet_name} (VLAN {vlan_id})')
    
    # Título
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f'{sheet_name.upper()} - VLAN {vlan_id} ({vlan_name})'
    title_cell.font = Font(bold=True, size=14, color=COLORS['header_text'])
    title_cell.fill = PatternFill(start_color=VLANS[vlan_id]['color'], end_color=VLANS[vlan_id]['color'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Informações da VLAN
    ws.merge_cells('A2:M2')
    info_cell = ws['A2']
    info_cell.value = f"Sub-rede: {VLANS[vlan_id]['subnet']} | Gateway: {VLANS[vlan_id]['gateway']} | {VLANS[vlan_id]['desc']}"
    info_cell.font = Font(size=10, italic=True, bold=True)
    info_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(info_cell)
    ws.row_dimensions[2].height = 20
    
    # Encontrar linha de cabeçalho
    header_row = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x) for x in row.values if pd.notna(x)]).upper()
        if 'ORDEM' in row_str and 'IP' in row_str:
            header_row = idx
            break
    
    if header_row is None:
        header_row = 4
    
    # Cabeçalhos
    original_headers = [str(x) for x in df.iloc[header_row].values]
    headers = ['VLAN', 'IP ANTIGO', 'IP NOVO', 'STATUS', 'ALERTAS'] + original_headers
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=10, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 50
    
    # Identificar colunas
    ip_col_idx = None
    tag_col_idx = None
    modelo_col_idx = None
    localizacao_col_idx = None
    
    for col_idx in range(len(df.columns)):
        col_header = str(df.iloc[header_row, col_idx]).upper()
        if 'IP' in col_header and ip_col_idx is None:
            ip_col_idx = col_idx
        if 'TAG' in col_header:
            tag_col_idx = col_idx
        if 'MODELO' in col_header:
            modelo_col_idx = col_idx
        if 'LOCALIZA' in col_header:
            localizacao_col_idx = col_idx
    
    # Processar dados
    row_num = 4
    ip_duplicates = []
    
    for idx in range(header_row + 1, len(df)):
        row_data = df.iloc[idx].values
        
        if pd.isna(row_data[0]) and (ip_col_idx is None or pd.isna(row_data[ip_col_idx])):
            continue
        
        ip_antigo = extract_ip(row_data[ip_col_idx]) if ip_col_idx is not None else None
        tag = row_data[tag_col_idx] if tag_col_idx is not None else None
        modelo = row_data[modelo_col_idx] if modelo_col_idx is not None else None
        localizacao = row_data[localizacao_col_idx] if localizacao_col_idx is not None else None
        
        # Verificar IPs duplicados
        alertas = []
        if ip_antigo and ip_antigo in assigned_ips:
            alertas.append("IP duplicado")
            ip_duplicates.append(ip_antigo)
        
        ip_novo = calculate_new_ip(ip_antigo, vlan_id, tag)
        status = 'Migrado' if ip_novo else 'Pendente'
        
        # Adicionar alertas de segurança
        if tag and ('OPERADOR' in str(tag).upper() or 'ADMIN' in str(tag).upper()):
            alertas.append("Credencial padrão")
        
        alerta_str = ', '.join(alertas) if alertas else 'OK'
        
        new_row = [f'VLAN {vlan_id}', ip_antigo, ip_novo, status, alerta_str] + list(row_data)
        ws.append(new_row)
        
        # Formatar linha
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx == 1:  # Coluna VLAN
                cell.fill = PatternFill(start_color=VLANS[vlan_id]['color'], end_color=VLANS[vlan_id]['color'], fill_type='solid')
            elif col_idx == 4:  # Coluna Status
                if status == 'Migrado':
                    cell.fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
            elif col_idx == 5:  # Coluna Alertas
                if 'duplicado' in alerta_str:
                    cell.fill = PatternFill(start_color=COLORS['error'], end_color=COLORS['error'], fill_type='solid')
                elif 'Credencial' in alerta_str:
                    cell.fill = PatternFill(start_color=COLORS['security'], end_color=COLORS['security'], fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        
        row_num += 1
    
    # Adicionar rodapé com estatísticas
    ws.merge_cells(f'A{row_num}:M{row_num}')
    stats_cell = ws.cell(row=row_num, column=1)
    stats_cell.value = f"Total de equipamentos: {row_num - 4} | IPs duplicados encontrados: {len(ip_duplicates)}"
    stats_cell.font = Font(bold=True, size=10)
    stats_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    apply_border(stats_cell)
    ws.row_dimensions[row_num].height = 20
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_migration_sheet(wb, excel_file):
    """Cria planilha de mapeamento de migração"""
    ws = wb.create_sheet('MAPEAMENTO MIGRACAO')
    
    # Título
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'MAPEAMENTO DE MIGRAÇÃO - IP ANTIGO → IP NOVO'
    title_cell.font = Font(bold=True, size=14, color=COLORS['header_text'])
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Cabeçalhos
    headers = ['IP ANTIGO', 'IP NOVO', 'VLAN ID', 'VLAN Nome', 'TAG Equipamento', 'Tipo', 'Observações']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[2].height = 40
    
    # Processar arquivo original
    xls = pd.ExcelFile(excel_file)
    row_num = 3
    
    # Processar CFTV
    if 'IP - CFTV' in xls.sheet_names:
        df = pd.read_excel(excel_file, sheet_name='IP - CFTV', header=None)
        # Adicionar mapeamentos (implementação simplificada)
        ws.append(['10.10.0.1', '10.10.40.51', 40, 'CFTV', 'CA-001-CP-01', 'Câmera', 'Migração CFTV'])
    
    # Ajustar largura das colunas
    column_widths = [15, 15, 12, 18, 20, 15, 30]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_management_sheet(wb):
    """Cria planilha de equipamentos de gerenciamento"""
    ws = wb.create_sheet('Equipamentos', 2)
    
    # Título
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'EQUIPAMENTOS DE GERENCIAMENTO - VLAN 10 (Management)'
    title_cell.font = Font(bold=True, size=14, color=COLORS['header_text'])
    title_cell.fill = PatternFill(start_color=VLANS[10]['color'], end_color=VLANS[10]['color'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    headers = ['VLAN ID', 'VLAN Nome', 'IP', 'TAG', 'Equipamento', 'Modelo', 'Localização', 'Função']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        apply_border(cell)
    ws.row_dimensions[2].height = 40
    
    switches = [
        [10, 'Management', '10.10.10.2', 'SW-00-RK-00', 'Switch Principal Rack', 'DS-3E1526P-EI', 'Rack Principal', 'Switch Core'],
        [10, 'Management', '10.10.10.3', 'SW-01-TR-01', 'Switch Torre 1', 'DS-3E1318P-EI/M', 'Torre 1', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.4', 'SW-02-TR-02', 'Switch Torre 2', 'DS-3E1526P-EI', 'Torre 2', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.5', 'SW-03-TR-03', 'Switch Torre 3', 'DS-3E1526P-EI', 'Torre 3', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.6', 'SW-04-TR-04', 'Switch Torre 4', 'DS-3E1526P-EI', 'Torre 4', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.7', 'SW-01-CP-01', 'Switch Perimetral 1', 'DS-3E1309P-EI', 'Caixa Perimetral 1', 'Switch Acesso'],
        [10, 'Management', '10.10.10.8', 'SW-02-CP-02', 'Switch Perimetral 2', 'DS-3E1309P-EI', 'Caixa Perimetral 2', 'Switch Acesso'],
        [10, 'Management', '10.10.10.9', 'SW-03-CP-03', 'Switch Perimetral 3', 'DS-3E1309P-EI', 'Caixa Perimetral 3', 'Switch Acesso'],
        [10, 'Management', '10.10.10.10', 'SW-04-CP-04', 'Switch Perimetral 4', 'DS-3E1309P-EI', 'Caixa Perimetral 4', 'Switch Acesso'],
        [10, 'Management', '10.10.10.1', 'ROTEADOR-01', 'Roteador/Gateway', 'MIKROTIK RB750Gr3', 'Rack Principal', 'Gateway'],
    ]
    
    for row_data in switches:
        ws.append(row_data)
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=VLANS[10]['color'], end_color=VLANS[10]['color'], fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_num].height = 25
    
    column_widths = [12, 18, 15, 15, 25, 20, 25, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_ip_addressing_sheet(wb):
    """Cria planilha de endereçamento IP detalhado"""
    ws = wb.create_sheet('Endereçamento IP', 3)
    
    # Título
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'DETALHAMENTO DO ENDEREÇAMENTO IP POR VLAN'
    title_cell.font = Font(bold=True, size=14, color=COLORS['header_text'])
    title_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(title_cell)
    ws.row_dimensions[1].height = 30
    
    # Informações gerais
    ws.merge_cells('A2:G2')
    info_cell = ws['A2']
    info_cell.value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de IPs atribuídos: {len(assigned_ips)}'
    info_cell.font = Font(size=10, italic=True)
    info_cell.fill = PatternFill(start_color=COLORS['info'], end_color=COLORS['info'], fill_type='solid')
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    apply_border(info_cell)
    ws.row_dimensions[2].height = 20
    
    # Cabeçalhos
    headers = ['VLAN ID', 'Nome VLAN', 'Sub-rede', 'Gateway', 'Faixa Utilizável', 'IPs Atribuídos', 'Utilização']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color=COLORS['header_text'])
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[3].height = 40
    
    # Dados das VLANs
    vlan_details = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', '10.10.10.2-254', '-', '0%'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', '10.10.20.2-254', '-', '0%'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', '10.10.30.2-254', '-', '0%'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', '10.10.40.2-254', '-', '0%'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', '10.10.50.2-254', '-', '0%'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', '10.10.60.2-254', '-', '0%'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', '10.10.100.2-254', '-', '0%']
    ]
    
    for row_idx, row_data in enumerate(vlan_details, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            vlan_id = row_data[0]
            cell.fill = PatternFill(start_color=VLANS[vlan_id]['color'], end_color=VLANS[vlan_id]['color'], fill_type='solid')
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            apply_border(cell)
        ws.row_dimensions[row_idx].height = 25
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 18, 15, 20, 15, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def main():
    """Função principal"""
    print("Gerando planilha Excel atualizada com VLANs...")
    
    # Resetar contadores e IPs atribuídos
    global ip_counters, assigned_ips
    ip_counters = {
        10: {'switches': 2, 'nvrs': 51, 'others': 100},
        40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
        50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
    }
    assigned_ips = set()
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    excel_file = 'Cópia de IP Senhas Condominio Calabasas.xlsx'
    xls = pd.ExcelFile(excel_file)
    
    # Criar planilhas
    create_summary_sheet(wb)
    
    if 'IP - CFTV' in xls.sheet_names:
        df_cftv = pd.read_excel(excel_file, sheet_name='IP - CFTV', header=None)
        process_equipment_sheet(wb, 'VLANs Detalhadas', df_cftv, 40, 'CFTV')
    
    if 'IP - CONTROLE DE ACESSO' in xls.sheet_names:
        df_access = pd.read_excel(excel_file, sheet_name='IP - CONTROLE DE ACESSO', header=None)
        process_equipment_sheet(wb, 'Controle de Acesso', df_access, 50, 'Access Control')
    
    create_management_sheet(wb)
    create_migration_sheet(wb, excel_file)
    create_ip_addressing_sheet(wb)
    
    output_file = 'Projeto_Logico_Rede_Calabasas_Atualizado.xlsx'
    wb.save(output_file)
    print(f"Planilha gerada com sucesso: {output_file}")

if __name__ == '__main__':
    main()