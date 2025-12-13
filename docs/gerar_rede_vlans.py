import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuração de cores para as VLANs
VLAN_COLORS = {
    10: 'FFE6E6',  # Vermelho claro - Management
    20: 'E6F3FF',  # Azul claro - Data
    30: 'E6FFE6',  # Verde claro - Voice
    40: 'FFF4E6',  # Laranja claro - CFTV
    50: 'F0E6FF',  # Roxo claro - Access Control
    60: 'FFFFE6',  # Amarelo claro - IoT
    100: 'E6E6E6'  # Cinza claro - Guest
}

# Mapeamento de VLANs
VLANS = {
    10: {'name': 'Management', 'subnet': '10.10.10.0/24', 'gateway': '10.10.10.1'},
    20: {'name': 'Data', 'subnet': '10.10.20.0/24', 'gateway': '10.10.20.1'},
    30: {'name': 'Voice', 'subnet': '10.10.30.0/24', 'gateway': '10.10.30.1'},
    40: {'name': 'CFTV', 'subnet': '10.10.40.0/24', 'gateway': '10.10.40.1'},
    50: {'name': 'Access Control', 'subnet': '10.10.50.0/24', 'gateway': '10.10.50.1'},
    60: {'name': 'IoT', 'subnet': '10.10.60.0/24', 'gateway': '10.10.60.1'},
    100: {'name': 'Guest', 'subnet': '10.10.100.0/24', 'gateway': '10.10.100.1'}
}

def extract_ip(ip_str):
    """Extrai IP de uma string"""
    if pd.isna(ip_str):
        return None
    ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    match = re.search(ip_pattern, str(ip_str))
    return match.group(0) if match else None

def get_vlan_for_equipment(tag, modelo, localizacao, ip_atual):
    """Determina a VLAN baseado no tipo de equipamento"""
    if pd.isna(tag):
        tag = ''
    if pd.isna(modelo):
        modelo = ''
    if pd.isna(localizacao):
        localizacao = ''
    
    tag_str = str(tag).upper()
    modelo_str = str(modelo).upper()
    localizacao_str = str(localizacao).upper()
    
    # Switches e roteadores - VLAN 10 (Management)
    if any(x in tag_str for x in ['SW-', 'RK-', 'ROTEADOR']):
        return 10
    
    # NVRs - VLAN 10 para gerenciamento, mas também podem ter interface na VLAN 40/50
    if 'NVR' in tag_str or 'NVR' in modelo_str:
        return 10
    
    # Câmeras CFTV - VLAN 40
    if 'CA-' in tag_str and any(x in localizacao_str for x in ['PERIMETRO', 'ESTACIONAMENTO', 'TORRE', 'AVENIDA']):
        return 40
    
    # Câmeras de elevador - VLAN 50
    if 'EL-' in tag_str or 'ELEVADOR' in localizacao_str:
        return 50
    
    # LPR - VLAN 50
    if 'LPR' in tag_str or 'LPR' in localizacao_str:
        return 50
    
    # Leitores biométricos - VLAN 50
    if 'LF-' in tag_str or 'LEITOR' in modelo_str:
        return 50
    
    # Controladores de acesso - VLAN 50
    if 'CE-' in tag_str or 'CONTROLADOR' in modelo_str:
        return 50
    
    # Câmeras de portaria - VLAN 50
    if 'PORTARIA' in localizacao_str or 'PORTÃO' in localizacao_str:
        return 50
    
    # Conversores de mídia - mesma VLAN do que conectam (geralmente CFTV)
    if 'CM-' in tag_str or 'CONVERSOR' in modelo_str:
        return 40  # Assumindo que são para CFTV
    
    # Por padrão, se IP está em 10.10.0.x, é CFTV (VLAN 40)
    if ip_atual and ip_atual.startswith('10.10.0.'):
        return 40
    
    # Se IP está em 10.10.1.x, é Controle de Acesso (VLAN 50)
    if ip_atual and ip_atual.startswith('10.10.1.'):
        return 50
    
    return 40  # Default para CFTV

# Contadores para IPs sequenciais por VLAN
ip_counters = {
    10: {'switches': 2, 'nvrs': 51, 'others': 100},
    40: {'cameras': 51, 'nvrs': 2, 'converters': 201},
    50: {'cameras_elevator': 11, 'cameras_portaria': 51, 'readers': 101, 'controllers': 201, 'nvrs': 2}
}

def calculate_new_ip(old_ip, vlan_id, equipment_type=None):
    """Calcula novo IP baseado na VLAN e tipo de equipamento"""
    global ip_counters
    
    if not old_ip:
        return None
    
    try:
        last_octet = int(old_ip.split('.')[-1])
        
        if vlan_id == 10:  # Management
            # Switches: 10.10.10.2-50
            # NVRs: 10.10.10.51-100
            if 'SW' in str(equipment_type).upper() or 'ROTEADOR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['switches']
                ip_counters[10]['switches'] += 1
                if ip_counters[10]['switches'] > 50:
                    ip_counters[10]['switches'] = 50
            elif 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[10]['nvrs']
                ip_counters[10]['nvrs'] += 1
                if ip_counters[10]['nvrs'] > 100:
                    ip_counters[10]['nvrs'] = 100
            else:
                new_octet = ip_counters[10]['others']
                ip_counters[10]['others'] += 1
            return f'10.10.10.{new_octet}'
        
        elif vlan_id == 40:  # CFTV
            # NVRs: 10.10.40.2-10
            # Câmeras: 10.10.40.51-200
            # Conversores: 10.10.40.201-240
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['nvrs']
                ip_counters[40]['nvrs'] += 1
                if ip_counters[40]['nvrs'] > 10:
                    ip_counters[40]['nvrs'] = 10
            elif 'CM-' in str(equipment_type).upper() or 'CONVERSOR' in str(equipment_type).upper():
                new_octet = ip_counters[40]['converters']
                ip_counters[40]['converters'] += 1
                if ip_counters[40]['converters'] > 240:
                    ip_counters[40]['converters'] = 240
            else:
                # Câmeras
                new_octet = ip_counters[40]['cameras']
                ip_counters[40]['cameras'] += 1
                if ip_counters[40]['cameras'] > 200:
                    ip_counters[40]['cameras'] = 200
            return f'10.10.40.{new_octet}'
        
        elif vlan_id == 50:  # Access Control
            # NVRs: 10.10.50.2-10
            # Câmeras elevador: 10.10.50.11-50
            # Câmeras portaria/LPR: 10.10.50.51-100
            # Leitores: 10.10.50.101-200
            # Controladores: 10.10.50.201-240
            if 'NVR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['nvrs']
                ip_counters[50]['nvrs'] += 1
            elif 'EL-' in str(equipment_type).upper() or 'ELEVADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['cameras_elevator']
                ip_counters[50]['cameras_elevator'] += 1
                if ip_counters[50]['cameras_elevator'] > 50:
                    ip_counters[50]['cameras_elevator'] = 50
            elif 'LF-' in str(equipment_type).upper() or 'LEITOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['readers']
                ip_counters[50]['readers'] += 1
                if ip_counters[50]['readers'] > 200:
                    ip_counters[50]['readers'] = 200
            elif 'CE-' in str(equipment_type).upper() or 'CONTROLADOR' in str(equipment_type).upper():
                new_octet = ip_counters[50]['controllers']
                ip_counters[50]['controllers'] += 1
                if ip_counters[50]['controllers'] > 240:
                    ip_counters[50]['controllers'] = 240
            else:
                # Câmeras portaria/LPR
                new_octet = ip_counters[50]['cameras_portaria']
                ip_counters[50]['cameras_portaria'] += 1
                if ip_counters[50]['cameras_portaria'] > 100:
                    ip_counters[50]['cameras_portaria'] = 100
            return f'10.10.50.{new_octet}'
        
        else:
            # Outras VLANs
            base = vlan_id * 10 if vlan_id < 100 else 100
            return f'10.10.{base}.{last_octet}'
    
    except Exception as e:
        print(f"Erro ao calcular IP: {e}")
        return None

def create_vlan_workbook():
    """Cria novo workbook com estrutura de VLANs"""
    wb = Workbook()
    
    # Remover planilha padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Ler arquivo Excel original
    excel_file = 'Cópia de IP Senhas Condominio Calabasas.xlsx'
    
    # Processar cada planilha do arquivo original
    xls = pd.ExcelFile(excel_file)
    
    # Criar planilha de resumo de VLANs
    ws_summary = wb.create_sheet('RESUMO VLANs', 0)
    create_summary_sheet(ws_summary)
    
    # Processar planilha IP - CFTV
    if 'IP - CFTV' in xls.sheet_names:
        df_cftv = pd.read_excel(excel_file, sheet_name='IP - CFTV', header=None)
        ws_cftv = wb.create_sheet('IP - CFTV (VLAN 40)')
        process_equipment_sheet(df_cftv, ws_cftv, 40)
    
    # Processar planilha IP - CONTROLE DE ACESSO
    if 'IP - CONTROLE DE ACESSO' in xls.sheet_names:
        df_access = pd.read_excel(excel_file, sheet_name='IP - CONTROLE DE ACESSO', header=None)
        ws_access = wb.create_sheet('IP - CONTROLE ACESSO (VLAN 50)')
        process_equipment_sheet(df_access, ws_access, 50)
    
    # Processar planilha PORTAS - CFTV
    if 'PORTAS - CFTV' in xls.sheet_names:
        df_ports_cftv = pd.read_excel(excel_file, sheet_name='PORTAS - CFTV', header=None)
        ws_ports_cftv = wb.create_sheet('PORTAS - CFTV (VLAN 40)')
        process_ports_sheet(df_ports_cftv, ws_ports_cftv, 40)
    
    # Processar planilha PORTAS - CONTROLE DE ACESSO
    if 'PORTAS - CONTROLE DE ACESSO' in xls.sheet_names:
        df_ports_access = pd.read_excel(excel_file, sheet_name='PORTAS - CONTROLE DE ACESSO', header=None)
        ws_ports_access = wb.create_sheet('PORTAS - CONTROLE ACESSO (VLAN 50)')
        process_ports_sheet(df_ports_access, ws_ports_access, 50)
    
    # Criar planilha de gerenciamento (VLAN 10)
    ws_mgmt = wb.create_sheet('GERENCIAMENTO (VLAN 10)')
    create_management_sheet(ws_mgmt)
    
    # Criar planilha de mapeamento de migração
    ws_migration = wb.create_sheet('MAPEAMENTO MIGRACAO')
    create_migration_mapping_sheet(ws_migration, excel_file)
    
    # Salvar arquivo
    output_file = 'Rede_VLANs_Condominio_Calabasas.xlsx'
    wb.save(output_file)
    print(f"Arquivo gerado: {output_file}")

def create_summary_sheet(ws):
    """Cria planilha de resumo de VLANs"""
    headers = ['VLAN ID', 'Nome', 'Sub-rede', 'Gateway', 'Máscara', 'Descrição', 'Faixa de IPs']
    ws.append(headers)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar dados das VLANs
    vlan_data = [
        [10, 'Management', '10.10.10.0/24', '10.10.10.1', '255.255.255.0', 'Gerenciamento de infraestrutura', '10.10.10.2-254'],
        [20, 'Data', '10.10.20.0/24', '10.10.20.1', '255.255.255.0', 'Rede corporativa', '10.10.20.2-254'],
        [30, 'Voice', '10.10.30.0/24', '10.10.30.1', '255.255.255.0', 'Telefonia IP', '10.10.30.2-254'],
        [40, 'CFTV', '10.10.40.0/24', '10.10.40.1', '255.255.255.0', 'Sistema de CFTV', '10.10.40.2-254'],
        [50, 'Access Control', '10.10.50.0/24', '10.10.50.1', '255.255.255.0', 'Controle de acesso', '10.10.50.2-254'],
        [60, 'IoT', '10.10.60.0/24', '10.10.60.1', '255.255.255.0', 'Dispositivos IoT', '10.10.60.2-254'],
        [100, 'Guest', '10.10.100.0/24', '10.10.100.1', '255.255.255.0', 'WiFi visitantes', '10.10.100.2-254']
    ]
    
    for row_idx, row_data in enumerate(vlan_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            vlan_id = row_data[0]
            if vlan_id in VLAN_COLORS:
                cell.fill = PatternFill(start_color=VLAN_COLORS[vlan_id], end_color=VLAN_COLORS[vlan_id], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 18, 15, 15, 30, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def process_equipment_sheet(df, ws, default_vlan):
    """Processa planilha de equipamentos e adiciona informações de VLAN"""
    # Encontrar linha de cabeçalho
    header_row = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x) for x in row.values if pd.notna(x)]).upper()
        if 'ORDEM' in row_str and 'IP' in row_str:
            header_row = idx
            break
    
    if header_row is None:
        header_row = 4  # Fallback
    
    # Copiar cabeçalhos e adicionar colunas de VLAN
    headers = ['VLAN ID', 'VLAN Nome', 'IP ANTIGO', 'IP NOVO'] + [str(x) for x in df.iloc[header_row].values]
    ws.append(headers)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Processar linhas de dados
    ip_col_idx = None
    tag_col_idx = None
    modelo_col_idx = None
    localizacao_col_idx = None
    
    # Identificar colunas
    for col_idx in range(len(df.columns)):
        col_header = str(df.iloc[header_row, col_idx]).upper()
        if 'IP' in col_header and ip_col_idx is None:
            ip_col_idx = col_idx
        if 'TAG' in col_header:
            tag_col_idx = col_idx
        if 'MODELO' in col_header:
            modelo_col_idx = col_idx
        if 'LOCALIZA' in col_header:
            localizacao_col_idx = col_idx
    
    row_num = 2
    for idx in range(header_row + 1, len(df)):
        row_data = df.iloc[idx].values
        
        # Verificar se é linha de dados válida
        if pd.isna(row_data[0]) and (ip_col_idx is None or pd.isna(row_data[ip_col_idx])):
            continue
        
        # Extrair informações
        ip_antigo = extract_ip(row_data[ip_col_idx]) if ip_col_idx is not None else None
        tag = row_data[tag_col_idx] if tag_col_idx is not None else None
        modelo = row_data[modelo_col_idx] if modelo_col_idx is not None else None
        localizacao = row_data[localizacao_col_idx] if localizacao_col_idx is not None else None
        
        # Determinar VLAN
        vlan_id = get_vlan_for_equipment(tag, modelo, localizacao, ip_antigo)
        if vlan_id is None:
            vlan_id = default_vlan
        
        vlan_name = VLANS[vlan_id]['name']
        
        # Calcular novo IP
        ip_novo = calculate_new_ip(ip_antigo, vlan_id, tag)
        
        # Criar linha
        new_row = [vlan_id, vlan_name, ip_antigo, ip_novo] + list(row_data)
        ws.append(new_row)
        
        # Formatar linha
        vlan_cell = ws.cell(row=row_num, column=1)
        if vlan_id in VLAN_COLORS:
            vlan_cell.fill = PatternFill(start_color=VLAN_COLORS[vlan_id], end_color=VLAN_COLORS[vlan_id], fill_type='solid')
        
        row_num += 1
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def process_ports_sheet(df, ws, default_vlan):
    """Processa planilha de portas"""
    # Similar ao process_equipment_sheet mas para portas
    headers = ['VLAN ID', 'VLAN Nome'] + [str(x) for x in df.iloc[3].values if pd.notna(x)]
    ws.append(headers)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Processar dados (simplificado)
    for idx in range(4, len(df)):
        row_data = df.iloc[idx].values
        new_row = [default_vlan, VLANS[default_vlan]['name']] + list(row_data)
        ws.append(new_row)
        
        # Formatar
        vlan_cell = ws.cell(row=idx-2, column=1)
        if default_vlan in VLAN_COLORS:
            vlan_cell.fill = PatternFill(start_color=VLAN_COLORS[default_vlan], end_color=VLAN_COLORS[default_vlan], fill_type='solid')

def create_management_sheet(ws):
    """Cria planilha de equipamentos de gerenciamento (VLAN 10)"""
    headers = ['VLAN ID', 'VLAN Nome', 'IP', 'TAG', 'Equipamento', 'Modelo', 'Localização', 'Função']
    ws.append(headers)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adicionar switches conhecidos
    switches = [
        [10, 'Management', '10.10.10.2', 'SW-00-RK-00', 'Switch Principal Rack', 'DS-3E1526P-EI', 'Rack Principal', 'Switch Core'],
        [10, 'Management', '10.10.10.3', 'SW-01-TR-01', 'Switch Torre 1', 'DS-3E1318P-EI/M', 'Torre 1', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.4', 'SW-02-TR-02', 'Switch Torre 2', 'DS-3E1526P-EI', 'Torre 2', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.5', 'SW-03-TR-03', 'Switch Torre 3', 'DS-3E1526P-EI', 'Torre 3', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.6', 'SW-04-TR-04', 'Switch Torre 4', 'DS-3E1526P-EI', 'Torre 4', 'Switch Distribuição'],
        [10, 'Management', '10.10.10.7', 'SW-01-CP-01', 'Switch Perimetral 1', 'DS-3E1309P-EI', 'Caixa Perimetral 1', 'Switch Acesso'],
        [10, 'Management', '10.10.10.8', 'SW-02-CP-02', 'Switch Perimetral 2', 'DS-3E1309P-EI', 'Caixa Perimetral 2', 'Switch Acesso'],
        [10, 'Management', '10.10.10.9', 'SW-03-CP-03', 'Switch Perimetral 3', 'DS-3E1309P-EI', 'Caixa Perimetral 3', 'Switch Acesso'],
        [10, 'Management', '10.10.10.10', 'SW-04-CP-04', 'Switch Perimetral 4', 'DS-3E1309P-EI', 'Caixa Perimetral 4', 'Switch Acesso'],
        [10, 'Management', '10.10.10.1', 'ROTEADOR-01', 'Roteador/Gateway', 'MIKROTIK RB750Gr3', 'Rack Principal', 'Gateway'],
    ]
    
    for row_data in switches:
        ws.append(row_data)
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=VLAN_COLORS[10], end_color=VLAN_COLORS[10], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar largura das colunas
    column_widths = [12, 18, 15, 15, 25, 20, 25, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_migration_mapping_sheet(ws, excel_file):
    """Cria planilha de mapeamento de migração IP antigo -> novo"""
    headers = ['IP ANTIGO', 'IP NOVO', 'VLAN ID', 'VLAN Nome', 'TAG Equipamento', 'Tipo', 'Observações']
    ws.append(headers)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Processar arquivo original para criar mapeamento
    xls = pd.ExcelFile(excel_file)
    
    # Processar CFTV
    if 'IP - CFTV' in xls.sheet_names:
        df = pd.read_excel(excel_file, sheet_name='IP - CFTV', header=None)
        # Adicionar mapeamentos CFTV
        # (implementação simplificada)
    
    # Ajustar largura das colunas
    column_widths = [15, 15, 12, 18, 20, 15, 30]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

if __name__ == '__main__':
    print("Gerando arquivo Excel com estrutura de VLANs...")
    create_vlan_workbook()
    print("Concluído!")

